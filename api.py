import os
import sys
import json
import re
import datetime
import logging
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from influxdb import InfluxDBClient
from dotenv import load_dotenv
import pytz
import jwt
import hmac
from functools import wraps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash

load_dotenv()

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

app = Flask(__name__, static_folder="dashboard/dist", static_url_path="")
CORS(app)

# SECRET_KEY es obligatoria: sin ella la app no arranca (evita el token JWT forjable).
# Genera una con: python -c "import secrets; print(secrets.token_hex(32))"
SECRET_KEY = os.getenv("SECRET_KEY", "")
if not SECRET_KEY:
    raise RuntimeError(
        "SECRET_KEY no estÃ¡ configurada en .env. "
        "Genera una con: python -c \"import secrets; print(secrets.token_hex(32))\""
    )


def hash_password(password: str) -> str:
    """Genera hash seguro (pbkdf2 con salt aleatorio)."""
    return generate_password_hash(password)


def verify_password(stored: str, password: str):
    """Verifica la contraseÃ±a contra lo almacenado.
    Retorna (ok, migrar_legacy): si almacenada en claro hay que re-guardarla hasheada.
    """
    if not stored or not isinstance(stored, str):
        return False, False
    try:
        if check_password_hash(stored, password):
            return True, False
    except Exception:
        pass  # formato viejo (texto plano) -> se valida abajo
    if hmac.compare_digest(stored, password or ""):
        return True, True  # legacy en claro, migrar a hash
    return False, False

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization")
        if not token:
            return jsonify({"error": "Token faltante"}), 401
        try:
            token = token.split(" ")[1] # Bearer <token>
            decoded = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            request.user = decoded
        except:
            return jsonify({"error": "Token invalido"}), 401
        return f(*args, **kwargs)
    return decorated

def role_required(role):
    def decorator(f):
        @wraps(f)
        @token_required
        def decorated(*args, **kwargs):
            if request.user.get("role") != role:
                return jsonify({"error": "Acceso restringido: se requiere rol de " + role}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


INFLUX_HOST = os.getenv("INFLUXDB_HOST", "influxdb")
INFLUX_PORT = int(os.getenv("INFLUXDB_PORT", 8086))
INFLUX_USER = os.getenv("INFLUXDB_ADMIN_USER", "admin")
INFLUX_PASSWORD = os.getenv("INFLUXDB_ADMIN_PASSWORD", "password")
INFLUX_DATABASE = os.getenv("INFLUXDB_DATABASE", "monitoreo")
INFLUX_RETENTION_DAYS = int(os.getenv("INFLUXDB_RETENTION_DAYS", "90"))
TRAFFIC_MEASUREMENT = f'"traffic_{INFLUX_RETENTION_DAYS}d"."traffic"'

# Modelos que son APs/infraestructura, no clientes finales
AP_MODELS = {"R5AC-Lite", "R5AC-Lite-2", "LAP-120", "LAP-GPS", "PBE-5AC", "PBE-M5", "NBE-5AC", "NBE-M5"}

# Nombres de mes en inglÃ©s (strftime %b) -> abreviatura en espaÃ±ol
MONTH_MAP = {
    "jan": "ene", "feb": "feb", "mar": "mar", "apr": "abr", "may": "may", "jun": "jun",
    "jul": "jul", "aug": "ago", "sep": "sep", "oct": "oct", "nov": "nov", "dec": "dic",
}


def get_influx_client():
    return InfluxDBClient(
        host=INFLUX_HOST,
        port=INFLUX_PORT,
        username=INFLUX_USER,
        password=INFLUX_PASSWORD,
        database=INFLUX_DATABASE,
        timeout=30,
    )


def query_latest_sites(client):
    """Obtiene el ultimo punto de cada site sin perder tags.
    El backend escribe cada 5 min, asi que el ultimo punto de cada serie
    siempre cae dentro de esta ventana; acotar por tiempo evita escanear
    todo el historico (retencion infinita) y los timeouts de lectura."""
    result = client.query("SELECT * FROM sites WHERE time > now() - 24h ORDER BY time DESC")
    latest = {}
    if result and "series" in getattr(result, "raw", {}):
        for series in result.raw["series"]:
            tags = series.get("tags", {})
            columns = series["columns"]
            for values in series.get("values", []):
                row = tags.copy()
                for col, val in zip(columns, values):
                    if col == "time":
                        # InfluxDB raw devuelve nanosegundos â†’ convertir a ISO string
                        if isinstance(val, (int, float)):
                            from datetime import datetime, timezone
                            row["time"] = datetime.fromtimestamp(val / 1e9, tz=timezone.utc).isoformat()
                        else:
                            row["time"] = val
                    elif val is None:
                        row[col] = 0
                    else:
                        row[col] = val
                sid = row.get("site_id", "")
                if sid and sid not in latest:
                    ls = row.get("last_seen")
                    if not ls or ls == 0 or not str(ls).strip():
                        row["last_seen"] = ""
                    latest[sid] = row

    # La fecha real de Ãºltima conexiÃ³n de cada junta: el device con lastSeen mÃ¡s
    # reciente de ese sitio (la API no expone lastSeen en el overview de /sites).
    _enrich_site_last_seen(client, latest)
    return list(latest.values())


def _enrich_site_last_seen(client, sites_map):
    """Completa last_seen de cada site con el Ãºltimo last_seen de sus devices."""
    try:
        dev_res = client.query("SELECT last_seen FROM devices WHERE time > now() - 24h GROUP BY site_id ORDER BY time DESC LIMIT 1")
        if not dev_res or "series" not in getattr(dev_res, "raw", {}):
            return
        for series in dev_res.raw["series"]:
            sid = (series.get("tags") or {}).get("site_id", "")
            site = sites_map.get(sid)
            if not site or site.get("last_seen"):
                continue
            values = series.get("values") or []
            if not values:
                continue
            col_idx = series["columns"].index("last_seen")
            ls = values[0][col_idx]
            if ls and str(ls).strip() and str(ls) != "0":
                site["last_seen"] = str(ls)
    except Exception as e:
        logger.warning(f"No se pudo enriquecer last_seen desde devices: {e}")


def query_latest_devices(client):
    """Obtiene el ultimo punto de cada device sin perder tags.
    El backend escribe cada 5 min, asi que el ultimo punto de cada serie
    siempre cae dentro de esta ventana; acotar por tiempo evita escanear
    todo el historico (retencion infinita) y los timeouts de lectura."""
    # Usamos SELECT * para obtener todos los campos originales sin el prefijo last_
    # Limitado por ventana de tiempo para rendimiento, agrupado por device_id
    result = client.query("SELECT * FROM devices WHERE time > now() - 24h GROUP BY device_id ORDER BY time DESC LIMIT 1")
    latest = []
    if result and "series" in getattr(result, "raw", {}):
        for series in result.raw["series"]:
            columns = series["columns"]
            values = series.get("values", [])[0]
            row = dict(zip(columns, values))
            
            # Asegurar que el ID venga de los tags si no estÃ¡ en columns
            tags = series.get("tags", {})
            if "device_id" not in row or row["device_id"] is None:
                row["device_id"] = tags.get("device_id")
            if "site_id" not in row or row["site_id"] is None:
                row["site_id"] = tags.get("site_id")
            latest.append(row)
    return latest


def query_site_devices_full(client, site_id):
    """Ultimo punto de CADA device del sitio en todo el historico (retencion
    infinita). La ventana de 24h oculta devices caidos hace mas tiempo, ya que
    el backend los escribe con timestamp = last_seen (fecha vieja)."""
    result = client.query(
        f"SELECT * FROM devices WHERE site_id = '{site_id}' GROUP BY device_id ORDER BY time DESC LIMIT 1"
    )
    latest = []
    if result and "series" in getattr(result, "raw", {}):
        for series in result.raw["series"]:
            columns = series["columns"]
            values = series.get("values", [])[0]
            row = dict(zip(columns, values))
            tags = series.get("tags", {})
            if "device_id" not in row or row["device_id"] is None:
                row["device_id"] = tags.get("device_id")
            if "site_id" not in row or row["site_id"] is None:
                row["site_id"] = tags.get("site_id")
            latest.append(row)
    return latest


def is_site_online(site):
    """Determina si un site esta online basado en estado y dispositivos"""
    status = site.get("status", "")
    if status == "active":
        return True
    if status in ("unknown",):
        return site.get("devices_available", 0) > 0
    return False


def get_site_health(site):
    """Determina la salud del sitio basada exclusivamente en sus Access Points"""
    ap_total = site.get("ap_count", 0)
    ap_online = site.get("ap_online", 0)
    
    if ap_total == 0:
        return "unknown"
    if ap_online == 0:
        return "total_caida"
    if ap_online >= ap_total:
        return "total_online"
    return "parcial"


def get_site_clients(site):
    """Clientes (hogares) de un site EXCLUYENDO los APs.
    El device_count de la API incluye los APs (66 = 64 clientes + 2 APs).
    Prioriza el conteo real de LB5 del backend; si no hay LB5 registrados,
    cae a device_count - ap_count."""
    lb5_total = int(site.get("lb5_count", 0) or 0)
    if lb5_total > 0:
        return lb5_total, int(site.get("lb5_online", 0) or 0)
    total = max(0, int(site.get("device_count", 0) or 0) - int(site.get("ap_count", 0) or 0))
    online = max(0, int(site.get("devices_available", 0) or 0) - int(site.get("ap_online", 0) or 0))
    return total, online


def get_override_estado(overrides, code):
    """Helper para obtener el estado desde los overrides, manejando strings o dicts"""
    val = overrides.get(code)
    if isinstance(val, dict):
        return val.get("estado", "")
    return val or ""

# ---------------------------------------------------------------------------
# ENDPOINTS
# ---------------------------------------------------------------------------

def load_users():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "users.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}

def save_users(users):
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "users.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(users, f, indent=2)

@app.route("/api/login", methods=["POST"])
def login():
    auth = request.get_json()
    username = auth.get("username")
    password = auth.get("password")
    
    users = load_users()
    user = users.get(username)
    
    if not user:
        return jsonify({"error": "Credenciales invÃ¡lidas"}), 401

    ok, migrate = verify_password(user.get("password"), password)
    if not ok:
        return jsonify({"error": "Credenciales invÃ¡lidas"}), 401

    # Migrar contraseÃ±a en claro (formato legacy) a hash
    if migrate:
        users[username]["password"] = hash_password(password)
        save_users(users)

    token = jwt.encode({
        "user": username, 
        "role": user.get("role", "viewer"), 
        "name": user.get("name", username),
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }, SECRET_KEY, algorithm="HS256")
    
    return jsonify({"token": token, "role": user.get("role", "viewer"), "name": user.get("name", username)})



@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "timestamp": datetime.datetime.now().isoformat()})


@app.route("/api/stats")
@token_required
def stats():
    """Estadisticas generales: APs online/offline, clientes online/offline"""
    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        devices = query_latest_devices(client)
        overrides = load_estado_overrides()

        total_aps = len(sites)
        aps_online = sum(1 for s in sites if is_site_online(s))
        aps_offline = total_aps - aps_online

        ops_total = 0
        ops_online = 0
        for s in sites:
            code_match = re.match(r"^(\d+)", s.get("site_name", "").strip())
            code = code_match.group(1) if code_match else ""
            estado = (get_override_estado(overrides, code) or s.get("estado", "") or "").strip()
            if estado.lower() == "operaciÃ³n":
                ops_total += 1
                if is_site_online(s):
                    ops_online += 1

        total_clients = len(devices)
        clients_online = sum(1 for d in devices if str(d.get("status", "")).lower() == "active")
        clients_offline = total_clients - clients_online

        total_download = sum(s.get("download_capacity", 0) for s in sites)
        total_upload = sum(s.get("upload_capacity", 0) for s in sites)
        avg_sla = (
            sum(s.get("sla", 0) for s in sites) / total_aps if total_aps > 0 else 0
        )

        return jsonify(
            {
                "aps_online": aps_online,
                "aps_offline": aps_offline,
                "aps_total": total_aps,
                "ops_online": ops_online,
                "ops_total": ops_total,
                "clients_online": clients_online,
                "clients_offline": clients_offline,
                "clients_total": total_clients,
                "total_download_mbps": round(total_download / 1e6, 2),
                "total_upload_mbps": round(total_upload / 1e6, 2),
                "avg_sla": round(avg_sla, 2),
            }
        )
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/sites")
@token_required
def sites_list():
    """Lista de APs con filtros: department, status, search, estado, health, grupo"""
    department = request.args.get("department")
    status = request.args.get("status")
    search = request.args.get("search", "").lower()
    estado = request.args.get("estado")
    health = request.args.get("health")
    grupo = request.args.get("grupo")
    
    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        sites = [s for s in sites if s.get("department") and s.get("department") not in ["unknown", "TPT"]]

        # CÃ³digos de juntas con ticket abierto (del Excel de tickets)
        open_by_code = get_tickets_open_map()

        # Aplicar overrides de estado
        overrides = load_estado_overrides()
        for s in sites:
            s["online"] = is_site_online(s)
            s["grupo"] = s.get("grupo") or "Sin grupo"
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            code = code_match.group(1) if code_match else ""
            tids = open_by_code.get(code, []) if code else []
            s["ticket_abierto"] = len(tids) > 0
            s["tickets_abiertos"] = len(tids)
            s["tickets_num"] = tids
            if code_match:
                if code in overrides:
                    s["estado"] = get_override_estado(overrides, code)
                # Health basado en APs (Rocket 5AC Lite)
            ap_total = s.get("ap_count", 0)
            ap_online = s.get("ap_online", 0)
            if ap_total == 0:
                s["health"] = "sin_aps"
            elif ap_online == 0:
                s["health"] = "caido"
            elif ap_online >= ap_total:
                s["health"] = "total"
            else:
                s["health"] = "parcial"

        if department:
            sites = [s for s in sites if s.get("department", "").lower() == department.lower()]
        if status:
            sites = [s for s in sites if s.get("status", "").lower() == status.lower()]
        if search:
            sites = [s for s in sites if search in s.get("site_name", "").lower()]
        if estado:
            sites = [s for s in sites if s.get("estado", "") == estado]
        if health and health != "all":
            sites = [s for s in sites if s.get("health") == health]
        if grupo:
            sites = [s for s in sites if s.get("grupo", "").lower() == grupo.lower()]

        return jsonify(sites)
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/traffic")
@token_required
def traffic():
    """Serie temporal de trafico Rx/Tx. Soporta ?site_id, ?hours, ?from, ?to, ?department, ?estado"""
    site_id = request.args.get("site_id")
    hours = request.args.get("hours")
    date_from = request.args.get("from")
    date_to = request.args.get("to")
    department = request.args.get("department")
    estado = request.args.get("estado")
    
    client = get_influx_client()
    try:
        if date_from and date_to:
            time_filter = f"time >= '{date_from}' AND time <= '{date_to}'"
        elif hours:
            time_filter = f"time > now() - {hours}h"
        else:
            time_filter = "time > now() - 24h"

        site_filter = f"AND site_id = '{site_id}'" if site_id else ""
        dept_filter = f"AND department = '{department}'" if department else ""
        estado_filter = f"AND estado = '{estado}'" if estado else ""

        query = f"""
            SELECT SUM(download_capacity) AS download,
                   SUM(upload_capacity) AS upload
            FROM {TRAFFIC_MEASUREMENT}
            WHERE {time_filter} {site_filter} {dept_filter} {estado_filter}
            GROUP BY TIME(5m)
        """
        result = client.query(query)
        points = []
        if result and "series" in getattr(result, "raw", {}):
            for series in result.raw["series"]:
                tags = series.get("tags", {})
                for value in series.get("values", []):
                    point = {
                        "time": value[0],
                        "download": value[1] if value[1] is not None else 0,
                        "upload": value[2] if value[2] is not None else 0,
                    }
                    if date_from:
                        point[site_id or "site"] = tags.get("site_id", "")
                    points.append(point)
        return jsonify(points)
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/devices")
@token_required
def devices_list():
    """Lista de dispositivos cliente con filtros"""
    site_id = request.args.get("site_id")
    status = request.args.get("status")
    search = request.args.get("search", "").lower()
    
    client = get_influx_client()
    try:
        if site_id:
            devices = query_site_devices_full(client, site_id)
        else:
            devices = query_latest_devices(client)
        
        if status:
            devices = [
                d for d in devices if str(d.get("status", "")).lower() == status.lower()
            ]
        if search:
            devices = [
                d
                for d in devices
                if search in str(d.get("device_name", "")).lower()
            ]
        
        return jsonify(devices)
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()




@app.route("/api/activity")
@token_required
def activity():
    """Actividad historica: clientes online/offline en el tiempo.
    Toma el ultimo punto de cada site por hora para evitar duplicados."""
    site_id = request.args.get("site_id")
    hours = request.args.get("hours")
    date_from = request.args.get("from")
    date_to = request.args.get("to")
    department = request.args.get("department")
    estado = request.args.get("estado")
    
    client = get_influx_client()
    try:
        site_filter = f"AND site_id = '{site_id}'" if site_id else ""
        dept_filter = f"AND department = '{department}'" if department else ""
        estado_filter = f"AND estado = '{estado}'" if estado else ""

        if date_from and date_to:
            time_filter = f"time >= '{date_from}' AND time <= '{date_to}'"
        elif hours:
            time_filter = f"time > now() - {hours}h"
        else:
            time_filter = "time > now() - 168h"

        query = f"""
            SELECT * FROM sites
            WHERE {time_filter} {site_filter} {dept_filter} {estado_filter}
            GROUP BY site_id
        """
        result = client.query(query)

        hour_site_latest = {}

        if result and "series" in getattr(result, "raw", {}):
            for series in result.raw["series"]:
                tags = series.get("tags", {})
                sid = tags.get("site_id", "")
                columns = series["columns"]

                try:
                    time_idx = columns.index("time")
                    dc_idx = columns.index("device_count")
                    da_idx = columns.index("devices_available")
                except ValueError:
                    continue

                for values in series.get("values", []):
                    ts = values[time_idx]
                    dc = values[dc_idx] if values[dc_idx] is not None else 0
                    da = values[da_idx] if values[da_idx] is not None else 0
                    hour_key = ts[:13] + ":00:00Z"
                    key = f"{hour_key}|{sid}"
                    hour_site_latest[key] = {
                        "hour": hour_key,
                        "total": int(float(dc)),
                        "online": int(float(da)),
                    }

        points = {}
        for key, val in hour_site_latest.items():
            h = val["hour"]
            if h not in points:
                points[h] = {"total": 0, "online": 0}
            points[h]["total"] += val["total"]
            points[h]["online"] += val["online"]

        result_list = sorted(
            [
                {
                    "time": h,
                    "total": v["total"],
                    "online": v["online"],
                    "offline": v["total"] - v["online"],
                }
                for h, v in points.items()
            ],
            key=lambda x: x["time"],
        )

        return jsonify(result_list)
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/site-devices")
@token_required
def site_devices():
    """Conteo de dispositivos por modelo para un site. Incluye APs y CPEs."""
    site_id = request.args.get("site_id", "")

    if not site_id:
        return jsonify({"error": "site_id requerido"}), 400

    client = get_influx_client()
    try:
        # Query all devices for the site. Since site_id is a field, we filter here.
        # We order by time DESC to get the most recent points first.
        # Acotado a ultimas 24h: el backend escribe cada 5 min, el ultimo punto
        # de cada device siempre cae dentro de esta ventana.
        query = f"SELECT * FROM devices WHERE site_id = '{site_id}' AND time > now() - 24h ORDER BY time DESC"
        result = client.query(query)
        
        latest = {}
        if result and "series" in getattr(result, "raw", {}):
            for series in result.raw["series"]:
                columns = series["columns"]
                for values in series.get("values", []):
                    row = dict(zip(columns, values))
                    did = row.get("device_id")
                    if did and did not in latest:
                        latest[did] = row



        lb5_total = 0
        lb5_online = 0
        ap_total = 0
        ap_online = 0

        for device in latest.values():
            model = str(device.get("device_model", "")).strip()
            status = str(device.get("status", "")).lower()
            is_online = status == "active"

            if model == "LB5":
                lb5_total += 1
                if is_online:
                    lb5_online += 1
            elif model in AP_MODELS:
                ap_total += 1
                if is_online:
                    ap_online += 1

        # Fallback para sitios caÃ­dos hace >24h: la ventana de devices ya no los
        # incluye, por lo que subestima el total. El conteo real de hogares sale
        # de la mediciÃ³n sites (lb5_count del backend).
        homes_site = lb5_total
        aps_site = ap_total
        try:
            site_res = client.query(
                f"SELECT * FROM sites WHERE site_id = '{site_id}' ORDER BY time DESC LIMIT 1"
            )
            if site_res and "series" in getattr(site_res, "raw", {}):
                for series in site_res.raw["series"]:
                    columns = series["columns"]
                    for values in series.get("values", []):
                        row = dict(zip(columns, values))
                        lb5c = int(row.get("lb5_count") or 0)
                        apc = int(row.get("ap_count") or 0)
                        if lb5c:
                            homes_site = lb5c
                        else:
                            dc = int(row.get("device_count") or 0)
                            homes_site = max(0, dc - apc)
                        aps_site = apc
                        break
        except Exception:
            pass

        lb5_total = max(lb5_total, homes_site)
        ap_total = max(ap_total, aps_site)

        return jsonify(
            {
                "site_id": site_id,
                "hogares_total": lb5_total,
                "hogares_online": lb5_online,
                "hogares_offline": lb5_total - lb5_online,
                "aps_total": ap_total,
                "aps_online": ap_online,
                "aps_offline": ap_total - ap_online,
            }
        )
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/hogares-bajos")
@token_required
def hogares_bajos():
    """Juntas en OPERACIÃ“N con menos del X% de hogares conectados (LB5 online/total).
    Filtros: pct (default 33), department (opcional)."""
    try:
        pct = float(request.args.get("pct", "33"))
    except ValueError:
        pct = 33.0
    department = request.args.get("department")

    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        overrides = load_estado_overrides()
        open_by_code = get_tickets_open_map()

        meta = {}
        for s in sites:
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            code = code_match.group(1) if code_match else ""
            estado = get_override_estado(overrides, code) or s.get("estado", "")
            if estado != "OperaciÃ³n":
                continue
            dept = s.get("department") or ""
            if dept in ("unknown", "TPT"):
                continue
            if department and department.lower() not in dept.lower():
                continue
            meta[s.get("site_id", "")] = {
                "site_id": s.get("site_id", ""),
                "site_name": s.get("site_name", ""),
                "department": dept,
                "last_online": s.get("last_seen") or "",
                "grupo": s.get("grupo") or "Sin grupo",
                "tickets_num": list(open_by_code.get(code, [])),
                "homes_site": int(s.get("lb5_count", 0) or 0) or max(
                    0,
                    int(s.get("device_count", 0) or 0) - int(s.get("ap_count", 0) or 0),
                ),
            }

        # Contar LB5 por sitio en una sola pasada (un query ya cacheado por device)
        devices = query_latest_devices(client)
        lb5_by_site = {}
        for d in devices:
            model = str(d.get("device_model", "")).strip()
            if model != "LB5":
                continue
            sid = d.get("site_id", "")
            if sid not in meta:
                continue
            st = str(d.get("status", "")).lower()
            info = lb5_by_site.setdefault(sid, {"total": 0, "online": 0})
            info["total"] += 1
            if st == "active":
                info["online"] += 1

        result = []
        for sid, m in meta.items():
            # Total de hogares = instalados/asignados al site (device_count - ap_count),
            # sin subestimar por la ventana de 24h de devices (juntas caÃ­das dejan de reportar).
            total = max(
                lb5_by_site.get(sid, {}).get("total", 0),
                m.get("homes_site", 0),
            )
            online = lb5_by_site.get(sid, {}).get("online", 0)
            if total == 0:
                continue
            pct_online = round(online / total * 100, 1)
            if pct_online < pct:
                result.append({
                    **m,
                    "hogares_total": total,
                    "hogares_online": online,
                    "hogares_offline": total - online,
                    "pct_online": pct_online,
                    "ths": 33,
                })

        result.sort(key=lambda r: r["pct_online"])
        return jsonify(result)
    except Exception as e:
        logger.exception("Error en /api/hogares-bajos")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/alerts")
@token_required
def alerts():
    """Sites con todos los dispositivos caidos, solo juntas en operacion"""
    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        offline_sites = []
        overrides = load_estado_overrides()
        open_by_code = get_tickets_open_map()
        for s in sites:
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            code = code_match.group(1) if code_match else ""
            estado = get_override_estado(overrides, code) or s.get("estado", "")
            if estado != "OperaciÃ³n":
                continue
            ap_total = s.get("ap_count", 0)
            ap_online = s.get("ap_online", 0)
            if ap_total > 0 and ap_online == 0:
                sid = s.get("site_id", "")
                # Buscar ultima vez que tuvo APs online
                last_online = None
                hours_down = None
                try:
                    q = f"SELECT ap_online FROM sites WHERE site_id = '{sid}' AND ap_online > 0 ORDER BY time DESC LIMIT 1"
                    r = client.query(q)
                    if r and "series" in getattr(r, "raw", {}) and r.raw["series"]:
                        last_ts = r.raw["series"][0]["values"][0][0]
                        last_online = last_ts
                        last_dt = datetime.fromisoformat(last_ts.replace("Z", "+00:00"))
                        now_utc = datetime.now(pytz.UTC)
                        hours_down = round((now_utc - last_dt).total_seconds() / 3600, 1)
                except Exception:
                    pass

                offline_sites.append(
                    {
                        "site_id": sid,
                        "site_name": s.get("site_name", ""),
                        "department": s.get("department", ""),
                        "ap_total": ap_total,
                        "ap_online": ap_online,
                        "device_count": s.get("device_count", 0),
                        "estado": estado,
                        "fecha_inicio": s.get("fecha_inicio", ""),
                        # Fecha real de Ãºltima conexiÃ³n: preferir el last_seen real
                        # del site (Ãºltimo device), fallback a la Ãºltima vez con APs online.
                        "last_online": s.get("last_seen") or last_online or "",
                        "hours_down": hours_down if hours_down is not None else "sin registro",
                        "grupo": s.get("grupo") or "Sin grupo",
                        # Info de tickets abiertos de la junta (para el modal)
                        "ticket_abierto": len(open_by_code.get(code, [])) > 0,
                        "tickets_abiertos": len(open_by_code.get(code, [])),
                        "tickets_num": open_by_code.get(code, []),
                    }
                )
        return jsonify(offline_sites)
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()


@app.route("/api/report")
@token_required
def report():
    """Reporte avanzado con estados de salud y filtros de operaciÃ³n"""
    dept_filter = request.args.get("department")
    type_filter = request.args.get("type") # 'operacion', 'implementacion', 'all'
    health_filter = request.args.get("health") # 'total_online', 'parcial', 'total_caida', 'all'
    search_filter = request.args.get("search", "").lower()
    
    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        overrides = load_estado_overrides()
        for s in sites:
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            if code_match:
                code = code_match.group(1)
                if code in overrides:
                    s["estado"] = get_override_estado(overrides, code)
  
        depts = {}
        for s in sites:
            dept = s.get("department") or "unknown"
            if dept in ["unknown", "TPT"]:
                continue
            
            # --- LÃ³gica de Salud del Sitio (Basada en APs/Rocket) ---
            clients_total, clients_online = get_site_clients(s)
            health_status = get_site_health(s)
            is_active = health_status != "total_caida" and health_status != "unknown"
            
            # --- LÃ³gica de Estado Operativo ---
            op_status = s.get("estado", "ImplementaciÃ³n").lower()
            
            # Aplicar Filtros
            if dept_filter and dept_filter.lower() not in dept.lower():
                continue
            if search_filter and search_filter not in s.get("site_name", "").lower():
                continue
            
            # Filtro de Tipo (OperaciÃ³n / ImplementaciÃ³n)
            if type_filter and type_filter != "all":
                if type_filter == "operacion" and op_status != "operaciÃ³n":
                    continue
                if type_filter == "implementacion" and op_status != "implementaciÃ³n":
                    continue

            # Filtro de Salud (Online / Parcial / CaÃ­da)
            if health_filter and health_filter != "all":
                if health_filter == "total_online" and health_status != "total_online":
                    continue
                if health_filter == "parcial" and health_status != "parcial":
                    continue
                if health_filter == "total_caida" and health_status != "total_caida":
                    continue
                
            key = dept
  
  
            if key not in depts:
                depts[key] = {
                    "department": dept,
                    "total_sites": 0,
                    "sites_online": 0,
                    "total_clients": 0,
                    "clients_online": 0,
                    "sites": [],
                }
  
            depts[key]["total_sites"] += 1
            if is_active:
                depts[key]["sites_online"] += 1
            
            depts[key]["total_clients"] += clients_total
            depts[key]["clients_online"] += clients_online
            
            last_seen_val = ""
            if s.get("last_seen"):
                last_seen_val = s["last_seen"]
            elif health_status != "total_caida":
                last_seen_val = s.get("time", "")

            depts[key]["sites"].append(
                {
                    "name": s.get("site_name", ""),
                    "status": s.get("status", ""),
                    "health": health_status,
                    "op_status": s.get("estado", "ImplementaciÃ³n"),
                    "online": health_status == "total_online",
                    "grupo": s.get("grupo") or "Sin grupo",
                    "clients": clients_total,
                    "clients_online": clients_online,
                    "ap_total": s.get("ap_count", 0),
                    "ap_online": s.get("ap_online", 0),
                    "download": round(s.get("download_capacity", 0) / 1e6, 4),
                    "upload": round(s.get("upload_capacity", 0) / 1e6, 4),
                    "last_seen": last_seen_val,
                    "lat": s.get("latitude", 0),
                    "lon": s.get("longitude", 0),

                }
            )
        
        result = sorted(depts.values(), key=lambda z: -z["total_clients"])
        return jsonify(result)
    except Exception as e:
        logger.exception("Error generando reporte")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()
 
 
@app.route("/api/report/excel/detailed")
@token_required
def report_excel_detailed():
    """Genera un reporte Excel detallado de todas las juntas con estilos"""
    dept_filter = request.args.get("department")
    type_filter = request.args.get("type")
    health_filter = request.args.get("health")
    search_filter = request.args.get("search", "").lower()
    
    client = get_influx_client()
    try:
        sites = query_latest_sites(client)
        overrides = load_estado_overrides()
        for s in sites:
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            if code_match:
                code = code_match.group(1)
                if code in overrides:
                    s["estado"] = get_override_estado(overrides, code)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detalle de Juntas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        headers = [
            "Junta", "Departamento", "Grupo", "TK Abiertos", 
            "Estado Salud", "Etapa Operativa", 
            "APs Totales", "APs Online", "Disponibilidad", "APs Offline", 
            "Clientes Totales", "Clientes Online", "Descarga (Mbps)", "Carga (Mbps)"
        ]
        ws.append(headers)
        
        # Estilo encabezado
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        open_by_code = get_tickets_open_map()

        included_sites = []
        for s in sites:
            dept = s.get("department") or "unknown"
            if dept in ["unknown", "TPT"]:
                continue
            
            clients_total, clients_online = get_site_clients(s)
            health_status = get_site_health(s)
            op_status = s.get("estado", "ImplementaciÃ³n")
            ap_total = s.get("ap_count", 0)
            ap_online = s.get("ap_online", 0)
            
            # Filtros (mismos que en /api/report)
            if dept_filter and dept_filter.lower() not in dept.lower(): continue
            if search_filter and search_filter not in s.get("site_name", "").lower(): continue
            if type_filter and type_filter != "all":
                if type_filter == "operacion" and op_status.lower() != "operaciÃ³n": continue
                if type_filter == "implementacion" and op_status.lower() != "implementaciÃ³n": continue
            if health_filter and health_filter != "all":
                if health_filter == "total_online" and health_status != "total_online": continue
                if health_filter == "parcial" and health_status != "parcial": continue
                if health_filter == "total_caida" and health_status != "total_caida": continue
                
            # Grupo y tickets abiertos de la junta
            code_match = re.match(r"^(\d+)", s.get("site_name", ""))
            code = code_match.group(1) if code_match else ""
            tids = open_by_code.get(code, []) if code else []
            tk_fmt = f"{len(tids)} Â· {', '.join(tids)}" if tids else "0"

            ws.append([
                s.get("site_name", ""),
                dept,
                s.get("grupo") or "Sin grupo",
                tk_fmt,
                health_status,
                op_status,
                ap_total,
                ap_online,
                f"{round((ap_online / ap_total * 100), 0)}%" if ap_total > 0 else "0%",
                ap_total - ap_online,
                clients_total,
                clients_online,
                round(s.get("download_capacity", 0) / 1e6, 4),
                round(s.get("upload_capacity", 0) / 1e6, 4)
            ])
            included_sites.append(s)

        # Hoja 2: APs de las juntas filtradas
        ws_aps = wb.create_sheet("APs")
        headers_aps = ["Junta", "Departamento", "Grupo", "Estado Salud", "AP Name", "Model", "Status", "IP", "MAC", "Signal", "Download", "Upload", "Last Connection"]
        ws_aps.append(headers_aps)
        for cell in ws_aps[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = center_alignment

        included_ids = {s.get("site_id") for s in included_sites}

        # Mapa site_id -> site (para nombre, dept, grupo, estado)
        site_by_id = {s.get("site_id"): s for s in included_sites}

        # HistÃ³rico completo por junta (incluye devices caÃ­dos hace >24h)
        all_devices = []
        for sid in included_ids:
            all_devices.extend(query_site_devices_full(client, sid))

        ap_rows = []
        for d in all_devices:
            sid = str(d.get("site_id", ""))
            if sid not in included_ids:
                continue
            if str(d.get("device_model", "")).strip() not in AP_MODELS:
                continue
            site = site_by_id[sid]
            last_conn = d.get("last_seen") or d.get("time")
            if last_conn:
                try:
                    dt = datetime.datetime.fromisoformat(str(last_conn).replace("Z", "+00:00"))
                    bogota_tz = pytz.timezone('America/Bogota')
                    last_conn_fmt = dt.astimezone(bogota_tz).strftime("%d/%m/%Y %H:%M:%S")
                except:
                    last_conn_fmt = str(last_conn)
            else:
                last_conn_fmt = "N/A"
            ap_rows.append([
                site.get("site_name", ""),
                site.get("department", ""),
                site.get("grupo") or "Sin grupo",
                get_site_health(site),
                d.get("device_name"),
                d.get("device_model"),
                d.get("status"),
                d.get("ip_address"),
                d.get("mac_address"),
                d.get("signal_strength"),
                d.get("rx_throughput"),
                d.get("tx_throughput"),
                last_conn_fmt
            ])

        ap_rows.sort(key=lambda r: (str(r[0] or ""), str(r[4] or "")))
        for row in ap_rows:
            ws_aps.append(row)

        # Hoja 3: Hogares (LB5) de las juntas filtradas
        ws_hog = wb.create_sheet("Hogares")
        headers_hog = ["Junta", "Departamento", "Grupo", "Estado Salud", "Hogar Name", "Model", "Status", "IP", "MAC", "Signal", "Download", "Upload", "Last Connection"]
        ws_hog.append(headers_hog)
        for cell in ws_hog[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = center_alignment

        hog_rows = []
        for d in all_devices:
            sid = str(d.get("site_id", ""))
            if sid not in included_ids:
                continue
            if str(d.get("device_model", "")).strip() != "LB5":
                continue
            site = site_by_id[sid]
            last_conn = d.get("last_seen") or d.get("time")
            if last_conn:
                try:
                    dt = datetime.datetime.fromisoformat(str(last_conn).replace("Z", "+00:00"))
                    bogota_tz = pytz.timezone('America/Bogota')
                    last_conn_fmt = dt.astimezone(bogota_tz).strftime("%d/%m/%Y %H:%M:%S")
                except:
                    last_conn_fmt = str(last_conn)
            else:
                last_conn_fmt = "N/A"
            hog_rows.append([
                site.get("site_name", ""),
                site.get("department", ""),
                site.get("grupo") or "Sin grupo",
                get_site_health(site),
                d.get("device_name"),
                d.get("device_model"),
                d.get("status"),
                d.get("ip_address"),
                d.get("mac_address"),
                d.get("signal_strength"),
                d.get("rx_throughput"),
                d.get("tx_throughput"),
                last_conn_fmt
            ])

        hog_rows.sort(key=lambda r: (str(r[0] or ""), str(r[4] or "")))
        for row in hog_rows:
            ws_hog.append(row)

        # Auto-ancho hoja Hogares
        for col in ws_hog.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws_hog.column_dimensions[column].width = max_length + 2
            
        # Auto-ancho columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Reporte_Detallado_{datetime.date.today()}.xlsx"
        )
    except Exception as e:
        logger.exception("Error generando reporte Excel detallado")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()

@app.route("/api/report/excel")

@token_required
def report_excel():
    """Genera un archivo Excel con la info de un site, sus APs y sus clientes"""
    site_id = request.args.get("site_id")
    if not site_id:
        return jsonify({"error": "site_id requerido"}), 400
 
    client = get_influx_client()
    try:
        # 1. Obtener info del site
        sites = query_latest_sites(client)
        site_data = next((s for s in sites if s.get("site_id") == site_id), None)
        if not site_data:
            return jsonify({"error": "Site no encontrado"}), 404
 
        # 2. Obtener dispositivos del site (histÃ³rico completo, incluye caÃ­dos >24h)
        devices = query_site_devices_full(client, site_id)
 
        # Crear Excel en memoria
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Hoja 1: Info General del Site
        ws1 = wb.active
        ws1.title = "InformaciÃ³n del Site"
        
        headers_site = ["Campo", "Valor"]
        ws1.append(headers_site)

        clients_total, clients_online = get_site_clients(site_data)
        
        site_info = [
            ("Site ID", site_data.get("site_id")),
            ("Site Name", site_data.get("site_name")),
            ("Departamento", site_data.get("department")),
            ("Estado Operativo", site_data.get("estado")),
            ("Estado Salud", get_site_health(site_data)),
            ("APs Totales", site_data.get("ap_count", 0)),
            ("APs Online", site_data.get("ap_online", 0)),
            ("Disponibilidad", f"{round((site_data.get('ap_online', 0) / site_data.get('ap_count', 1) * 100), 0)}%" if site_data.get("ap_count", 0) > 0 else "0%"),
            ("APs Offline", site_data.get("ap_count", 0) - site_data.get("ap_online", 0)),
            ("Clientes Totales", clients_total),
            ("Clientes Online", clients_online),
            ("Capacidad Download (Mbps)", round(site_data.get("download_capacity", 0) / 1e6, 4)),
            ("Capacidad Upload (Mbps)", round(site_data.get("upload_capacity", 0) / 1e6, 4)),
        ]
        for row in site_info:
            ws1.append(row)
            
        # Estilos Hoja 1
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Hoja 2: APs del Site
        ws_aps = wb.create_sheet("APs")
        headers_aps = ["AP ID", "AP Name", "Model", "Status", "IP", "MAC", "Signal", "Download", "Upload", "Last Connection"]
        ws_aps.append(headers_aps)
        
        # Hoja 3: Clientes del Site
        ws_clients = wb.create_sheet("Clientes")
        headers_clients = ["Client ID", "Client Name", "Model", "Status", "IP", "MAC", "Signal", "Download", "Upload", "Last Connection"]
        ws_clients.append(headers_clients)
 
        for d in devices:
            # Formatear fecha
            last_conn = d.get("last_seen") or d.get("time")
            if last_conn:
                try:
                    # InfluxDB dates are often in ISO format
                    dt = datetime.datetime.fromisoformat(last_conn.replace("Z", "+00:00"))
                    bogota_tz = pytz.timezone('America/Bogota')
                    dt_bogota = dt.astimezone(bogota_tz)
                    last_conn_fmt = dt_bogota.strftime("%d/%m/%Y %H:%M:%S")
                except:
                    last_conn_fmt = last_conn
            else:
                last_conn_fmt = "N/A"

            row = [
                d.get("device_id"),
                d.get("device_name"),
                d.get("device_model"),
                d.get("status"),
                d.get("ip_address"),
                d.get("mac_address"),
                d.get("signal_strength"),
                d.get("rx_throughput"),
                d.get("tx_throughput"),
                last_conn_fmt
            ]
            if str(d.get("device_model", "")).strip() in AP_MODELS:
                ws_aps.append(row)
            else:
                ws_clients.append(row)
        
        # Estilos Hoja 2 y 3
        for sheet in [ws_aps, ws_clients]:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Ajustar ancho de columnas automÃ¡ticamente (aproximado)
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 2
        
        # Eliminar hoja por defecto si existe
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.active = ws1
 
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Sanitizar nombre del archivo eliminando caracteres invÃ¡lidos
        raw_name = site_data.get('site_name', site_id)
        safe_name = re.sub(r'[\\/*?:"<>|]', "_", str(raw_name))
        filename = f"Reporte_{safe_name}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception("Error generando Excel")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()
 
 
@app.route("/api/toggle-estado", methods=["POST"])
@role_required("admin")
def toggle_estado():
    """Cambia el estado de una junta entre Operacion e Implementacion"""
    data = request.get_json() or {}
    code = data.get("code", "").strip()
    nuevo_estado = data.get("estado", "").strip()

    if not code or not nuevo_estado:
        return jsonify({"error": "code y estado requeridos"}), 400

    overrides = load_estado_overrides()
    fecha_inicio = data.get("fecha_inicio", "")
    overrides[code] = {"estado": nuevo_estado, "fecha_inicio": fecha_inicio}
    save_estado_overrides(overrides)
    
    return jsonify({"ok": True, "code": code, "estado": nuevo_estado, "fecha_inicio": fecha_inicio})


def get_estado_overrides_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "estado_overrides.json")


def load_estado_overrides():
    path = get_estado_overrides_path()
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def save_estado_overrides(data):
    path = get_estado_overrides_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f)


@app.route("/api/users/<username>", methods=["DELETE"])
@role_required("admin")
def delete_user(username):
    """Admin: Eliminar usuario"""
    users = load_users()
    if username not in users:
        return jsonify({"error": "Usuario no encontrado"}), 404
    
    # Evitar que el admin se elimine a si mismo (opcional pero recomendado)
    if request.user.get("user") == username:
        return jsonify({"error": "No puedes eliminarte a ti mismo"}), 400
        
    del users[username]
    save_users(users)
    return jsonify({"ok": True, "message": f"Usuario {username} eliminado"})

@app.route("/api/users", methods=["GET", "POST"])
@role_required("admin")
def manage_users():
    """Admin: Obtener o crear usuarios"""
    if request.method == "GET":
        users = load_users()
        # No enviar passwords en la respuesta
        safe_users = {k: {v: val for v, val in u.items() if v != "password"} for k, u in users.items()}
        return jsonify(safe_users)
    
    if request.method == "POST":
        data = request.get_json()
        username = data.get("username")
        password = data.get("password")
        role = data.get("role", "viewer")
        name = data.get("name", username)
        
        if not username or not password:
            return jsonify({"error": "Usuario y contraseÃ±a requeridos"}), 400
            
        users = load_users()
        users[username] = {"password": hash_password(password), "role": role, "name": name}
        save_users(users)
        return jsonify({"ok": True, "message": f"Usuario {username} creado/actualizado"})

@app.route("/api/departments")
@token_required
def departments():
    """Departamentos disponibles para filtros"""
    client = get_influx_client()
    try:
        result = client.query("SHOW TAG VALUES FROM sites WITH KEY = department")
        depts = []
        if result and "series" in getattr(result, "raw", {}):
            for series in result.raw["series"]:
                for value in series.get("values", []):
                    if value[1] and value[1] not in ["unknown", "TPT"]:
                        depts.append(value[1])
        return jsonify(sorted(set(depts)))
    except Exception as e:
        logger.exception("Error en la solicitud")
        return jsonify({"error": str(e)}), 500
    finally:
        client.close()






def get_sites_in_operacion():
    """Obtiene el mapeo de nombres de sites en OperaciÃ³n -> fecha_inicio"""
    client = get_influx_client()
    try:
        # 1. Cargar desde base_operacion.xlsx (Fuente maestra de operaciÃ³n)
        op_sites = {}
        op_excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel", "base_operacion.xlsx")
        if os.path.exists(op_excel_path):
            wb = openpyxl.load_workbook(op_excel_path, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 15 or not row[1]: continue
                code = str(row[1]).strip()
                # Col 7 (index 6) es el nombre real de la junta
                name = str(row[6] or "").strip().upper()
                if not name: name = code
                
                # Col 15 (index 14) es fecha inicio operaciÃ³n
                f_inicio = str(row[14] or "") if row[14] else ""
                op_sites[name] = {"fecha_inicio": f_inicio}

        
        # 2. Cargar overrides
        overrides = load_estado_overrides()
        
        # 3. Sincronizar con InfluxDB para capturar cambios dinÃ¡micos y nombres reales
        sites_influx = query_latest_sites(client)
        for s in sites_influx:
            full_name = s.get("site_name", "").strip()
            code_match = re.match(r"^(\d+)", full_name)
            code = code_match.group(1) if code_match else ""
            estado = (get_override_estado(overrides, code) or s.get("estado", "") or "").strip()
            
            if estado.lower() == "operaciÃ³n":
                name = full_name
                if code_match:
                    name = full_name[len(code):].strip()
                
                if name:
                    name_up = str(name).strip().upper()
                    f_inicio = ""
                    override_val = overrides.get(code)
                    if isinstance(override_val, dict):
                        f_inicio = override_val.get("fecha_inicio", "")
                    if not f_inicio:
                        f_inicio = s.get("fecha_inicio", "")
                    if not f_inicio and name_up in op_sites:
                        f_inicio = op_sites[name_up].get("fecha_inicio", "")
                    op_sites[name_up] = {"fecha_inicio": f_inicio}

        return op_sites
    except Exception as e:
        logger.exception("Error obteniendo sites en operacion")
        return {}

def get_preventivos_data(filters):
    """Lee el archivo Excel de preventivos y agrega los datos segÃºn filtros"""
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "Indicador preventivos V2.xlsx")
    if not os.path.exists(file_path):
        return {"error": "Archivo de preventivos no encontrado"}, 404
    
    try:
        op_sites = get_sites_in_operacion()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        month_order = {
            "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
            "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12
        }
        
        dept_filter = filters.get("department")
        month_filter = filters.get("month")
        year_filter = filters.get("year")

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        filtered_rows = []
        
        for row in rows:
            if not row or len(row) < 4: continue
            
            site_name = str(row[2] or "").strip().upper()
            if site_name not in op_sites:
                continue
            
            dept = str(row[3] or "").strip()
            
            # Fecha Preventivo (Index 15)
            date_val = row[15] if len(row) > 15 and isinstance(row[15], datetime.datetime) else None
            
            # Filtro de departamento
            if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower():
                continue
            
            # Filtros de tiempo
            if date_val:
                year = date_val.year
                month_name = date_val.strftime("%b").lower()
                MONTH_MAP
                month_es = month_map.get(month_name, month_name)
                
                if year_filter and str(year) != year_filter:
                    continue
                if month_filter and month_filter != "all" and month_filter != month_es:
                    continue
            else:
                if year_filter or (month_filter and month_filter != "all"):
                    continue
            
            filtered_rows.append(row)
        
        # --- LÃ³gica para la primera grÃ¡fica (Inicio OperaciÃ³n) ---
        COL_OP_START = 9
        max_year, max_month_val = 2025, 12
        
        for row in rows:
            if not row or len(row) <= COL_OP_START: continue
            site_name = str(row[2] or "").strip().upper()
            if site_name not in op_sites: continue
            dept = str(row[3] or "").strip()
            if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower(): continue
            
            op_date = row[COL_OP_START]
            if isinstance(op_date, datetime.datetime):
                op_year = op_date.year
                op_month_name = op_date.strftime("%b").lower()
                MONTH_MAP
                op_month_es = month_map.get(op_month_name, op_month_name)
                m_val = month_order.get(op_month_es, 0)
                if op_year > max_year or (op_year == max_year and m_val > max_month_val):
                    max_year, max_month_val = op_year, m_val
        
        stats = {}
        for row in rows:
            if not row or len(row) <= COL_OP_START: continue
            site_name = str(row[2] or "").strip().upper()
            if site_name not in op_sites: continue
            dept = str(row[3] or "").strip()
            if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower(): continue
            
            op_date = row[COL_OP_START]
            if isinstance(op_date, datetime.datetime):
                op_year = op_date.year
                op_month_name = op_date.strftime("%b").lower()
                MONTH_MAP
                op_month_es = month_map.get(op_month_name, op_month_name)
                key = (op_year, op_month_es)
                if key not in stats:
                    stats[key] = {"Ejecutado": 0, "Pendiente": 0}
                
                # Estado ProgramaciÃ³n Preventivo (Index 14, columna real de datos)
                estado_val = str(row[14] or "").strip() if len(row) > 14 else ""
                if "EJECUTADO" in estado_val.upper():
                    stats[key]["Ejecutado"] += 1
                else:
                    stats[key]["Pendiente"] += 1
        
        chart_data = []
        temp_year, temp_month_val = 2025, 12
        month_list = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
        while temp_year < max_year or (temp_year == max_year and temp_month_val <= max_month_val):
            m_name = month_list[temp_month_val - 1]
            key = (temp_year, m_name)
            val = stats.get(key, {"Ejecutado": 0, "Pendiente": 0})
            chart_data.append({"month": f"{m_name} {temp_year}", "Ejecutado": val["Ejecutado"], "Pendiente": val["Pendiente"]})
            temp_month_val += 1
            if temp_month_val > 12:
                temp_month_val = 1
                temp_year += 1
        
        exec_month_stats = {}
        for row in rows:
            if not row or len(row) < 3: continue
            site_name = str(row[2] or "").strip().upper()
            if site_name not in op_sites: continue
            dept = str(row[3] or "").strip()
            if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower(): continue
            
            # Usar la columna de estado (14 = Estado ProgramaciÃ³n Preventivo)
            estado_val = str(row[14] or "").strip() if len(row) > 14 else ""
            if "EJECUTADO" in estado_val.upper():
                date_val = row[15] if len(row) > 15 and isinstance(row[15], datetime.datetime) else None
                if date_val:
                    month_name = date_val.strftime("%b").lower()
                    MONTH_MAP
                    month_es = month_map.get(month_name, month_name)
                    if month_es:
                        exec_month_stats[month_es] = exec_month_stats.get(month_es, 0) + 1
        
        op_start_data = []
        for month_es in ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]:
            op_start_data.append({"month": month_es.capitalize(), "count": exec_month_stats.get(month_es, 0)})
            
        return {"chart_data": chart_data, "op_start_data": op_start_data, "total_filtered": len(filtered_rows)}, 200
    except Exception as e:
        logger.exception("Error leyendo preventivos")
        return {"error": str(e)}, 500


@app.route("/api/preventivos")
@token_required
def preventivos():
    """Datos para el indicador de mantenimientos preventivos"""
    filters = {
        "department": request.args.get("department"),
        "month": request.args.get("month"),
        "year": request.args.get("year"),
    }
    data, status_code = get_preventivos_data(filters)
    return jsonify(data), status_code

@app.route("/api/preventivos/excel")
@token_required
def preventivos_excel():
    """Exporta la lista de preventivos filtrada a Excel con estilos profesionales"""
    dept_filter = request.args.get("department")
    month_filter = request.args.get("month")
    year_filter = request.args.get("year")
    status_filter = request.args.get("status")
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "Indicador preventivos V2.xlsx")
    
    if not os.path.exists(file_path):
        return jsonify({"error": "Archivo no encontrado"}), 404
        
    try:
        op_sites = get_sites_in_operacion()
        wb_src = openpyxl.load_workbook(file_path, data_only=True)
        sheet_src = wb_src.active
        
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Preventivos Filtrados"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        date_alignment = Alignment(horizontal="center")
        
        headers = [cell.value for cell in sheet_src[1]]
        ws_out.append(headers)
        
        # Aplicar estilos al encabezado
        for cell in ws_out[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        for row in sheet_src.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3: continue
            # Filtro: Solo juntas en OPERACIÃ“N
            site_name = str(row[2] or "").strip().upper()
            if site_name not in op_sites:
                continue
            
            dept = str(row[3] or "").strip()
            date_val = row[15] if len(row) > 15 and isinstance(row[15], datetime.datetime) else (row[16] if len(row) > 16 and isinstance(row[16], datetime.datetime) else None)
            
            if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower():
                continue
            
            if isinstance(date_val, datetime.datetime):
                year = date_val.year
                month_name = date_val.strftime("%b").lower()
                MONTH_MAP
                month_es = month_map.get(month_name, month_name)
                
                if year_filter and str(year) != year_filter:
                    continue
                if month_filter and month_filter != "all" and month_filter != month_es:
                    continue
            else:
                if year_filter or (month_filter and month_filter != "all"):
                    continue
            
            if status_filter and status_filter != "all":
                estado_val = str(row[14] or "").strip() if len(row) > 14 else ""
                is_ejecutado = "EJECUTADO" in estado_val.upper()
                if status_filter == "ejecutado" and not is_ejecutado:
                    continue
                if status_filter == "pendiente" and is_ejecutado:
                    continue
            
            ws_out.append(row)
            
        # Estilos a las filas y ajuste de columnas
        for row in ws_out.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Centrar columna de fecha y estado (ajustar indices segÃºn headers)
                # Basado en el excel original: Col 15 (Estado) y Col 16 (Fecha)
                if cell.column == 15 or cell.column == 16:
                    cell.alignment = center_alignment

        # Auto-ancho columnas
        for col in ws_out.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws_out.column_dimensions[column].width = max_length + 4
            
        output = BytesIO()
        wb_out.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Reporte_Preventivos_{datetime.date.today()}.xlsx"
        )
    except Exception as e:
        logger.exception("Error exportando preventivos")
        return jsonify({"error": str(e)}), 500

@app.route("/api/preventivos/list")
@token_required
def preventivos_list():
    """Obtiene la lista completa de preventivos para la tabla de gestiÃ³n"""
    dept_filter = request.args.get("department")
    status_filter = request.args.get("status")
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "Indicador preventivos V2.xlsx")
    
    if not os.path.exists(file_path):
        return jsonify({"error": "Archivo no encontrado"}), 404
    try:
        op_sites = get_sites_in_operacion()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = []
        if sheet.max_row >= 1:
            headers = [cell.value for cell in sheet[1] if cell.value]
        data = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3: continue
            try:
                # indices 0-based: 1: ID JUNTA, 2: JUNTA DE INTERNET, 3: DEPARTAMENTO, 9: Fecha Inicio OperaciÃ³n, 
                # 14: Estado ProgramaciÃ³n Preventivo (data real), 15: Fecha Preventivo
                
                site_name = str(row[2] or "").strip().upper()
                if site_name not in op_sites:
                    continue
                
                dept = str(row[3] or "").strip()
                if dept_filter and dept_filter != "all" and dept_filter.lower() not in dept.lower():
                    continue
                
                row_dict = {}
                row_dict["ID_JUNTA"] = str(row[1] or "") if len(row) > 1 else ""
                row_dict["JUNTA"] = str(row[2] or "") if len(row) > 2 else ""
                row_dict["DEPARTAMENTO"] = dept if dept else "N/A"
                
                # Fecha Preventivo (Index 15)
                fecha_prev = ""
                if len(row) > 15 and isinstance(row[15], datetime.datetime):
                    fecha_prev = row[15].isoformat()
                row_dict["FECHA_PREVENTIVO"] = fecha_prev

                # Estado ProgramaciÃ³n Preventivo (Index 14, columna real de datos)
                estado_val = str(row[14] or "").strip() if len(row) > 14 else ""
                if not estado_val:
                    # Fallback: bÃºsqueda en toda la fila
                    for cell in row:
                        if cell and isinstance(cell, str):
                            if "EJECUTADO" in cell.upper():
                                estado_val = "Ejecutado"
                                break
                            elif "PENDIENTE" in cell.upper():
                                estado_val = "Pendiente"
                
                # Normalizar a "Ejecutado" o "Pendiente"
                if "EJECUTADO" in estado_val.upper():
                    final_estado = "Ejecutado"
                elif "PENDIENTE" in estado_val.upper() or not estado_val:
                    final_estado = "Pendiente"
                else:
                    final_estado = "Pendiente"
                
                row_dict["ESTADO_PREVENTIVO"] = final_estado

                if status_filter and status_filter != "all":
                    if status_filter == "ejecutado" and final_estado != "Ejecutado":
                        continue
                    if status_filter == "pendiente" and final_estado != "Pendiente":
                        continue
                
                for i, val in enumerate(row):
                    header = headers[i] if i < len(headers) else f"col_{i}"
                    if header:
                        norm_header = str(header).strip().upper()
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            row_dict[norm_header] = val.isoformat()
                        elif isinstance(val, datetime.time):
                            row_dict[norm_header] = val.strftime("%H:%M:%S")
                        else:
                            row_dict[norm_header] = val
                
                row_dict["inicio_operacion"] = op_sites.get(site_name, {}).get("fecha_inicio", "")
                
                data.append(row_dict)
            except Exception as row_e:
                logger.error(f"Error processing row {row}: {row_e}")
                continue
        
        return jsonify(data)
    except Exception as e:
        logger.exception("Error listando preventivos")
        return jsonify({"error": str(e)}), 500


@app.route("/api/preventivos/update", methods=["POST"])
@role_required("admin")
def update_preventivo():
    """Actualiza el estado y/o fecha de un preventivo en el Excel"""
    data = request.get_json() or {}
    # Usamos ID JUNTA como identificador
    junta_id = data.get("junta_id")
    new_status = data.get("status") # Ahora es opcional
    new_date_str = data.get("date")
    
    if not junta_id:
        return jsonify({"error": "junta_id requerido"}), 400
        
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "Indicador preventivos V2.xlsx")
    
    try:
        # Cargamos el libro sin data_only para poder escribir
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Columnas: 1: ID JUNTA, 15: Estado ProgramaciÃ³n Preventivo, 16: Fecha Preventivo
        # Nota: openpyxl usa indexacion 1. Col 22 (Estado Preventivo) es FORMULA derivada de col 15.
        COL_ID = 2 
        COL_DATE = 16
        COL_STATUS = 15
        
        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[COL_ID-1].value) == str(junta_id):
                # Actualizar estado si se proporcionÃ³
                if new_status:
                    row[COL_STATUS-1].value = new_status
                
                # Actualizar fecha si se proporcionÃ³
                if new_date_str:
                    try:
                        # Formato YYYY-MM-DD
                        date_val = datetime.datetime.strptime(new_date_str, "%Y-%m-%d")
                        row[COL_DATE-1].value = date_val
                    except ValueError:
                        return jsonify({"error": "Formato de fecha invalido. Use YYYY-MM-DD"}), 400
                
                found = True
                break
        
        if not found:
            return jsonify({"error": f"No se encontro la junta con ID {junta_id}"}), 404
            
        wb.save(file_path)
        return jsonify({"ok": True, "message": "Preventivo actualizado correctamente"})
    except Exception as e:
        logger.exception("Error actualizando preventivo")
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------------------------
# REPORTES DE TICKETS (Solo proyecto JUNTAS)
# ---------------------------------------------------------------------------

TICKET_COL = "#Ticket"
PARADA_DAYS_COL = "DÃ­as Parada Reloj"


def find_tickets_file():
    """Busca el archivo de tickets: prioridad a data/, luego la raiz del proyecto"""
    base = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base, "data", "tickets_juntas.xlsx"),
    ]
    for folder in [os.path.join(base, "data"), base]:
        if os.path.isdir(folder):
            for f in sorted(os.listdir(folder)):
                if f.startswith("DetalleTickets_") and f.endswith(".xlsx"):
                    candidates.append(os.path.join(folder, f))
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


_tickets_cache = {"mtime": None, "open_by_code": {}}


def get_tickets_open_map():
    """{codigo INRED (6 dÃ­gitos): [ticket_id abiertos]}"""
    file_path = find_tickets_file()
    if not file_path:
        return {}

    try:
        mtime = os.path.getmtime(file_path)
    except OSError:
        return {}

    if _tickets_cache["mtime"] == mtime:
        return _tickets_cache["open_by_code"]

    by_code = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            if not headers or "Ticket Estado" not in headers or "CÃ³digo Operador" not in headers:
                continue
            st_idx = headers.index("Ticket Estado")
            co_idx = headers.index("CÃ³digo Operador")
            tk_idx = headers.index(TICKET_COL) if TICKET_COL in headers else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or st_idx >= len(row) or co_idx >= len(row):
                    continue
                estado = str(row[st_idx] or "").strip().lower()
                if estado != "abierto":
                    continue
                m = re.search(r"(\d{6})", str(row[co_idx] or ""))
                if not m:
                    continue
                tid = None
                if tk_idx is not None and tk_idx < len(row) and row[tk_idx] is not None:
                    tid = str(row[tk_idx]).strip()
                by_code.setdefault(m.group(1), [])
                if tid:
                    by_code[m.group(1)].append(tid)
        wb.close()
    except Exception as e:
        logger.exception("Error cargando tickets abiertos")

    _tickets_cache["mtime"] = mtime
    _tickets_cache["open_by_code"] = by_code
    logger.info(f"Juntas con tickets abiertos: {len(by_code)}")
    return by_code


def load_tickets_data():
    """Carga los tickets del proyecto JUNTAS y las paradas de reloj del Excel.
    Retorna (tickets, paradas_por_ticket, file_path). Las hojas se detectan
    por sus columnas para tolerar archivos con nombres de hoja distintos."""
    file_path = find_tickets_file()
    if not file_path:
        return None, None, None

    wb = openpyxl.load_workbook(file_path, data_only=True)
    tickets = []
    paradas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if not headers:
            continue

        if TICKET_COL in headers:
            idx = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                rec = {}
                for h, i in idx.items():
                    rec[h] = row[i] if i < len(row) else None
                proyecto = str(rec.get("Proyecto") or "").strip().upper()
                if "JUNTA" not in proyecto:
                    continue
                tickets.append(rec)

        elif PARADA_DAYS_COL in headers:
            idx = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                rec = {}
                for h, i in idx.items():
                    rec[h] = row[i] if i < len(row) else None
                tno = str(rec.get("Ticket #") or "").strip()
                if tno:
                    paradas.setdefault(tno, []).append(rec)

    return tickets, paradas, file_path


def parse_dt(value):
    """Convierte un valor a datetime o None si no es fecha"""
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    if isinstance(value, str) and value.strip():
        try:
            return datetime.datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            try:
                return datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
    return None


def fmt_fecha_legible(value):
    """Formatea a dd/mm/aaaa hh:mm (legible en reportes) o cadena vacÃ­a"""
    dt = parse_dt(value)
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


def enrich_ticket(rec, paradas, now_dt):
    """Calcula horas totales, horas de parada de reloj y horas netas de un ticket.
    Los tickets con prioridad Baja no generan horas de indisponibilidad."""
    tno = str(rec.get("#Ticket") or "").strip()
    f_inicio = parse_dt(rec.get("Fecha Inicio"))
    f_fin = parse_dt(rec.get("Ticket Fecha Fin"))
    estado = str(rec.get("Ticket Estado") or "").strip()
    prioridad = str(rec.get("Prioridad") or "").strip()

    horas_total = 0.0
    if f_inicio:
        fin_dt = f_fin if f_fin else now_dt
        horas_total = max(0.0, (fin_dt - f_inicio).total_seconds() / 3600.0)

    dias_parada = 0.0
    for p in paradas.get(tno, []):
        try:
            dias_parada += float(p.get(PARADA_DAYS_COL) or 0)
        except (TypeError, ValueError):
            continue
    horas_parada = dias_parada * 24.0
    horas_netas = max(0.0, horas_total - horas_parada)

    es_prioridad_baja = "baja" in prioridad.lower()
    indisp_bruta = horas_total if not es_prioridad_baja else 0.0
    indisp_neta = horas_netas if not es_prioridad_baja else 0.0

    sub_proy = str(rec.get("Sub Proyecto") or "").strip().upper()
    if "SEGUIMIENTO INTERNO" in sub_proy:
        subtipo = "seguimiento_interno"
        subtipo_label = "Juntas seguimiento interno"
    else:
        subtipo = "juntas"
        subtipo_label = "Juntas"

    out = dict(rec)
    out.update({
        "ticket_estado": estado,
        "fecha_inicio": f_inicio.isoformat() if f_inicio else "",
        "fecha_fin": f_fin.isoformat() if f_fin else "",
        "horas_total": round(horas_total, 2),
        "dias_parada": round(dias_parada, 4),
        "horas_parada": round(horas_parada, 2),
        "horas_netas": round(horas_netas, 2),
        "prioridad_baja": es_prioridad_baja,
        "indisp_bruta_horas": round(indisp_bruta, 2),
        "indisp_neta_horas": round(indisp_neta, 2),
        "categoria": str(rec.get("Categoria") or "").strip(),
        "departamento": str(rec.get("Departamento") or "").strip(),
        "codigo_operador": str(rec.get("CÃ³digo Operador") or "").strip(),
        "municipio": str(rec.get("Municipio") or "").strip(),
        "centro_poblado": str(rec.get("Centro Poblado") or "").strip(),
        "tipo": str(rec.get("Tipo") or "").strip(),
        "responsable": str(rec.get("Responsable") or "").strip(),
        "prioridad": prioridad,
        "grupo_escalamiento": str(rec.get("Grupo Escalamiento") or "").strip(),
        "sub_proyecto": str(rec.get("Sub Proyecto") or "").strip(),
        "subtipo": subtipo,
        "subtipo_label": subtipo_label,
    })
    return out


def filter_tickets(tickets, filters, now_dt):
    """Aplica filtros comunes (department, junta, estado, tipo, from, to, search)"""
    out = []
    for t in tickets:
        if filters.get("department") and t.get("departamento", "").lower() != filters["department"].lower():
            continue
        if filters.get("junta") and filters["junta"] != "all":
            m = re.search(r"(\d{6})", t.get("codigo_operador", ""))
            code = m.group(1) if m else ""
            if code != filters["junta"]:
                continue
        if filters.get("estado") and filters["estado"] != "all":
            if t.get("ticket_estado", "").lower() != filters["estado"].lower():
                continue
        if filters.get("tipo") and filters["tipo"] != "all":
            if filters["tipo"].lower() == "incidente" and "incidente" not in t.get("tipo", "").lower():
                continue
            if filters["tipo"].lower() == "peticion" and "petici" not in t.get("tipo", "").lower():
                continue
        if filters.get("subtipo") and filters["subtipo"] != "all":
            if t.get("subtipo", "") != filters["subtipo"]:
                continue
        f_inicio = parse_dt(t.get("Fecha Inicio"))
        if filters.get("from"):
            f_from = parse_dt(filters["from"])
            if f_inicio and f_from and f_inicio < f_from:
                continue
        if filters.get("to"):
            f_to = parse_dt(filters["to"])
            if f_inicio and f_to and f_inicio > f_to:
                continue
        if filters.get("search"):
            s = filters["search"].lower()
            haystack = " ".join([
                t.get("ticket_estado", ""),
                str(t.get("#Ticket") or ""),
                t.get("codigo_operador", ""),
                t.get("departamento", ""),
                t.get("municipio", ""),
                t.get("centro_poblado", ""),
                t.get("categoria", ""),
            ]).lower()
            if s not in haystack:
                continue
        out.append(t)
    return out


@app.route("/api/tickets/stats")
@token_required
def tickets_stats():
    """KPIs de tickets del proyecto JUNTAS con filtros"""
    filters = {
        "department": request.args.get("department"),
        "junta": request.args.get("junta"),
        "estado": request.args.get("estado"),
        "tipo": request.args.get("tipo"),
        "subtipo": request.args.get("subtipo"),
        "from": request.args.get("from"),
        "to": request.args.get("to"),
        "search": request.args.get("search", "").strip(),
    }
    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado. Ubicelo como data/tickets_juntas.xlsx o DetalleTickets_*.xlsx en la raiz"}), 404

    now_dt = datetime.datetime.now()
    enriched = [enrich_ticket(t, paradas, now_dt) for t in tickets]
    filtered = filter_tickets(enriched, filters, now_dt)

    stats = {
        "total": len(filtered),
        "abiertos": 0,
        "cerrados": 0,
        "anulados": 0,
        "incidentes": 0,
        "peticiones": 0,
        "mttr_horas": 0.0,
        "indisponibilidad_bruta_horas": 0.0,
        "indisponibilidad_neta_horas": 0.0,
        "horas_parada": 0.0,
        "horas_prioridad_baja": 0.0,
        "tickets_prioridad_baja": 0,
        "tickets_con_parada": 0,
        "departments": sorted({t.get("departamento", "") for t in enriched if t.get("departamento")}),
    }

    mttr_sum = 0.0
    mttr_count = 0
    for t in filtered:
        estado = t.get("ticket_estado", "").lower()
        if estado == "abierto":
            stats["abiertos"] += 1
        elif estado == "cerrado":
            stats["cerrados"] += 1
            mttr_sum += t.get("horas_total", 0.0)
            mttr_count += 1
        elif estado == "anulado":
            stats["anulados"] += 1

        if "incidente" in t.get("tipo", "").lower():
            stats["incidentes"] += 1
        elif "petici" in t.get("tipo", "").lower():
            stats["peticiones"] += 1

        if t.get("horas_parada", 0) > 0:
            stats["tickets_con_parada"] += 1

        if t.get("prioridad_baja"):
            stats["tickets_prioridad_baja"] += 1
            stats["horas_prioridad_baja"] += t.get("horas_total", 0.0)

        stats["indisponibilidad_bruta_horas"] += t.get("indisp_bruta_horas", 0.0)
        stats["indisponibilidad_neta_horas"] += t.get("indisp_neta_horas", 0.0)
        stats["horas_parada"] += t.get("horas_parada", 0.0)

    if mttr_count > 0:
        stats["mttr_horas"] = round(mttr_sum / mttr_count, 2)
    stats["indisponibilidad_bruta_horas"] = round(stats["indisponibilidad_bruta_horas"], 2)
    stats["indisponibilidad_neta_horas"] = round(stats["indisponibilidad_neta_horas"], 2)
    stats["horas_parada"] = round(stats["horas_parada"], 2)
    stats["horas_prioridad_baja"] = round(stats["horas_prioridad_baja"], 2)
    return jsonify(stats)


@app.route("/api/tickets")
@token_required
def tickets_list():
    """Lista detallada de tickets JUNTAS con filtros y paginacion"""
    filters = {
        "department": request.args.get("department"),
        "junta": request.args.get("junta"),
        "estado": request.args.get("estado"),
        "tipo": request.args.get("tipo"),
        "subtipo": request.args.get("subtipo"),
        "from": request.args.get("from"),
        "to": request.args.get("to"),
        "search": request.args.get("search", "").strip(),
    }
    try:
        page = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(200, int(request.args.get("page_size", 50))))
    except ValueError:
        page, page_size = 1, 50

    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado"}), 404

    now_dt = datetime.datetime.now()
    enriched = [enrich_ticket(t, paradas, now_dt) for t in tickets]
    filtered = filter_tickets(enriched, filters, now_dt)
    filtered.sort(key=lambda t: t.get("fecha_inicio", ""), reverse=True)

    total = len(filtered)
    start = (page - 1) * page_size
    items = filtered[start:start + page_size]
    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "source": os.path.basename(file_path) if file_path else "",
    })


@app.route("/api/tickets/months")
@token_required
def tickets_months():
    """Tendencia mensual de tickets JUNTAS (creados, cerrados, abiertos)"""
    filters = {
        "department": request.args.get("department"),
        "junta": request.args.get("junta"),
        "estado": request.args.get("estado"),
        "tipo": request.args.get("tipo"),
        "subtipo": request.args.get("subtipo"),
        "from": request.args.get("from"),
        "to": request.args.get("to"),
        "search": request.args.get("search", "").strip(),
    }
    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado"}), 404

    now_dt = datetime.datetime.now()
    enriched = [enrich_ticket(t, paradas, now_dt) for t in tickets]
    filtered = filter_tickets(enriched, filters, now_dt)

    monthly = {}
    for t in filtered:
        f_inicio = parse_dt(t.get("Fecha Inicio"))
        if not f_inicio:
            continue
        key = f_inicio.strftime("%Y-%m")
        if key not in monthly:
            monthly[key] = {"month": key, "total": 0, "cerrados": 0, "abiertos": 0}
        monthly[key]["total"] += 1
        if t.get("ticket_estado", "").lower() == "cerrado":
            monthly[key]["cerrados"] += 1
        elif t.get("ticket_estado", "").lower() == "abierto":
            monthly[key]["abiertos"] += 1

    return jsonify(sorted(monthly.values(), key=lambda m: m["month"]))


@app.route("/api/tickets/excel")
@token_required
def tickets_excel():
    """Exporta los tickets JUNTAS filtrados a Excel con paradas de reloj aplicadas"""
    filters = {
        "department": request.args.get("department"),
        "junta": request.args.get("junta"),
        "estado": request.args.get("estado"),
        "tipo": request.args.get("tipo"),
        "subtipo": request.args.get("subtipo"),
        "from": request.args.get("from"),
        "to": request.args.get("to"),
        "search": request.args.get("search", "").strip(),
    }
    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado"}), 404

    now_dt = datetime.datetime.now()
    enriched = [enrich_ticket(t, paradas, now_dt) for t in tickets]
    filtered = filter_tickets(enriched, filters, now_dt)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Tickets JUNTAS"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "#Ticket", "CÃ³digo Operador", "Departamento", "Municipio", "Centro Poblado",
        "Fecha Inicio", "Fecha Fin", "Estado", "CategorÃ­a", "Sub Proyecto", "Tipo",
        "Grupo Escalamiento", "Prioridad", "Responsable",
        "Horas Total", "DÃ­as Parada Reloj", "Horas Parada", "Horas Netas",
        "Indisponibilidad Bruta (h)", "Indisponibilidad Neta (h)",
    ]
    ws_out.append(headers)
    for cell in ws_out[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for t in filtered:
        ws_out.append([
            t.get("#Ticket"),
            t.get("codigo_operador", ""),
            t.get("departamento", ""),
            t.get("municipio", ""),
            t.get("centro_poblado", ""),
            fmt_fecha_legible(t.get("fecha_inicio")),
            fmt_fecha_legible(t.get("fecha_fin")),
            t.get("ticket_estado", ""),
            t.get("categoria", ""),
            t.get("sub_proyecto", ""),
            t.get("tipo", ""),
            t.get("grupo_escalamiento", ""),
            t.get("prioridad", ""),
            t.get("responsable", ""),
            t.get("horas_total", 0),
            t.get("dias_parada", 0),
            t.get("horas_parada", 0),
            t.get("horas_netas", 0),
            t.get("indisp_bruta_horas", 0),
            t.get("indisp_neta_horas", 0),
        ])

    for col in ws_out.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_out.column_dimensions[column].width = min(max_length + 4, 60)

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Reporte_Tickets_JUNTAS_{datetime.date.today()}.xlsx",
    )


@app.route("/api/tickets/upload", methods=["POST"])
@role_required("admin")
def tickets_upload():
    """Carga un archivo Excel de tickets (hoja de tickets + hoja de paradas)"""
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No se recibiÃ³ ningÃºn archivo"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "El archivo debe ser .xlsx"}), 400

    dest = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "tickets_juntas.xlsx")
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    tmp_path = dest + ".tmp.xlsx"
    file.save(tmp_path)

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        has_tickets = False
        has_paradas = False
        for sheet_name in wb.sheetnames:
            headers = [str(c.value).strip() if c.value else "" for c in wb[sheet_name][1]]
            if TICKET_COL in headers:
                has_tickets = True
            if PARADA_DAYS_COL in headers:
                has_paradas = True
        if not has_tickets or not has_paradas:
            os.remove(tmp_path)
            return jsonify({"error": "El archivo no tiene las hojas esperadas (tickets y paradas de reloj)"}), 400
    except Exception as e:
        os.remove(tmp_path)
        return jsonify({"error": f"No se pudo leer el archivo: {e}"}), 400

    os.replace(tmp_path, dest)
    return jsonify({"ok": True, "message": "Archivo de tickets cargado correctamente"})


def parse_seguimiento(text):
    """Extrae las entradas de seguimiento de los comentarios del ticket"""
    if not text:
        return []
    out = []
    pattern = re.compile(
        r"TIPO Comentario:\s*([^=]*?)\s*==>\s*FECHA:([\d\- :.]+)\s*Comentario\s*:==>\s*(.*?)(?=TIPO Comentario:|$)",
        re.DOTALL,
    )
    for m in pattern.finditer(str(text)):
        out.append({
            "tipo": m.group(1).strip(),
            "fecha": m.group(2).strip(),
            "texto": m.group(3).strip().replace("######******######", "").strip(),
        })
    return out


@app.route("/api/tickets/<ticket_id>")
@token_required
def tickets_detail(ticket_id):
    """Detalle completo de un ticket con su seguimiento y paradas de reloj"""
    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado"}), 404

    ticket = next(
        (t for t in tickets if str(t.get("#Ticket") or "").strip() == str(ticket_id).strip()),
        None,
    )
    if not ticket:
        return jsonify({"error": f"Ticket {ticket_id} no encontrado"}), 404

    now_dt = datetime.datetime.now()
    enriched = enrich_ticket(ticket, paradas, now_dt)

    seguimiento = []
    seguimiento.extend(parse_seguimiento(ticket.get("Comentario Apertura")))
    seguimiento.extend(parse_seguimiento(ticket.get("Comentario SoluciÃ³n")))
    if not seguimiento:
        seguimiento.append({
            "tipo": "Apertura",
            "fecha": enriched.get("fecha_inicio", ""),
            "texto": str(ticket.get("Comentario Apertura") or "Sin comentarios").strip(),
        })

    tno = str(ticket_id).strip()
    paradas_ticket = []
    for p in paradas.get(tno, []):
        paradas_ticket.append({
            "fecha_inicio_falla": parse_dt(p.get("Fecha Inicio Falla")),
            "fecha_fin_falla": parse_dt(p.get("Fecha Fin Falla")),
            "fecha_inicio_parada": parse_dt(p.get("Fecha Inicio Parada")),
            "fecha_fin_parada": parse_dt(p.get("Fecha Fin Parada")),
            "dias_parada": p.get(PARADA_DAYS_COL),
        })
    for p in paradas_ticket:
        for k, v in p.items():
            if isinstance(v, datetime.datetime):
                p[k] = v.isoformat()
            elif v is None:
                p[k] = ""

    enriched["seguimiento"] = seguimiento
    enriched["paradas"] = paradas_ticket
    return jsonify(enriched)


def get_sites_coords():
    """Mapa codigo junta (6 digitos) -> lat/lon desde InfluxDB"""
    client = get_influx_client()
    coords = {}
    try:
        sites = query_latest_sites(client)
        for s in sites:
            name = s.get("site_name", "") or ""
            m = re.match(r"^(\d{6})", name.strip())
            code = m.group(1) if m else ""
            if code and s.get("latitude") and s.get("longitude"):
                coords[code] = {
                    "lat": s.get("latitude"),
                    "lon": s.get("longitude"),
                    "name": name.strip(),
                    "department": s.get("department", ""),
                }
    except Exception as e:
        logger.warning(f"Error obteniendo coordenadas de sites: {e}")
    finally:
        client.close()
    return coords


@app.route("/api/tickets/juntas")
@token_required
def tickets_juntas():
    """Agregacion por junta: total TK, indisponibilidad neta (excluye prioridad Baja) y coordenadas"""
    filters = {
        "department": request.args.get("department"),
        "junta": request.args.get("junta"),
        "estado": request.args.get("estado"),
        "tipo": request.args.get("tipo"),
        "subtipo": request.args.get("subtipo"),
        "from": request.args.get("from"),
        "to": request.args.get("to"),
        "search": request.args.get("search", "").strip(),
    }
    tickets, paradas, file_path = load_tickets_data()
    if tickets is None:
        return jsonify({"error": "Archivo de tickets no encontrado"}), 404

    now_dt = datetime.datetime.now()
    enriched = [enrich_ticket(t, paradas, now_dt) for t in tickets]
    filtered = filter_tickets(enriched, filters, now_dt)
    coords = get_sites_coords()

    juntas = {}
    for t in filtered:
        m = re.search(r"(\d{6})", t.get("codigo_operador", ""))
        code = m.group(1) if m else "SIN_CODIGO"
        if code not in juntas:
            coord = coords.get(code, {})
            juntas[code] = {
                "code": code,
                "name": coord.get("name", code),
                "department": coord.get("department", "") or t.get("departamento", ""),
                "lat": coord.get("lat", 0),
                "lon": coord.get("lon", 0),
                "total": 0,
                "abiertos": 0,
                "cerrados": 0,
                "incidentes": 0,
                "indisp_bruta_horas": 0.0,
                "indisp_neta_horas": 0.0,
                "horas_parada": 0.0,
            }
        j = juntas[code]
        j["total"] += 1
        estado = t.get("ticket_estado", "").lower()
        if estado == "abierto":
            j["abiertos"] += 1
        elif estado == "cerrado":
            j["cerrados"] += 1
        if "incidente" in t.get("tipo", "").lower():
            j["incidentes"] += 1
        j["indisp_bruta_horas"] += t.get("indisp_bruta_horas", 0.0)
        j["indisp_neta_horas"] += t.get("indisp_neta_horas", 0.0)
        j["horas_parada"] += t.get("horas_parada", 0.0)

    result = sorted(
        juntas.values(),
        key=lambda j: (j["indisp_neta_horas"], j["total"]),
        reverse=True,
    )
    for j in result:
        j["indisp_bruta_horas"] = round(j["indisp_bruta_horas"], 2)
        j["indisp_neta_horas"] = round(j["indisp_neta_horas"], 2)
        j["horas_parada"] = round(j["horas_parada"], 2)
    return jsonify(result)

# ---------------------------------------------------------------------------
# SERVIR REACT BUILD
# ---------------------------------------------------------------------------

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path):
    if path and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, "index.html")


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if not os.path.exists(os.path.join(app.static_folder, "index.html")):
        print(
            "WARNING: React build no encontrado. Ejecuta: cd dashboard && npm run build",
            file=sys.stderr,
        )

    port = int(os.getenv("DASHBOARD_PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
