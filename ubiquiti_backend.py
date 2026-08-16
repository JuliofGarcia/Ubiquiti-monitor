import requests
import time
import os
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from influxdb import InfluxDBClient
import pytz
from config import CONFIG as GlobalConfig

# CONFIGURACIÓN DE LOGGING
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ubiquiti_monitor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Retención aplicada en InfluxDB (días). Los points más antiguos que esto se
# descartan del batch para no provocar "partial write: points beyond retention".
RETENTION_DAYS = int(os.getenv('INFLUXDB_RETENTION_DAYS', '90'))


# ============================================================================
# CARGA DE MAPEO DESDE EXCEL
# ============================================================================
def load_excel_mapping(filepath: str = "excel/base_instalaciones.xlsx") -> Dict:
    """Carga el archivo Excel de instalaciones y genera mapeo por codigo INRED.
    Retorna dict: codigo_inred -> {department, centro_poblado, lat, lon}
    """
    mapping = {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        
        # Saltamos el encabezado y usamos índices
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            
            code = str(row[1]).strip()
            # Indices basados en el análisis de df.columns:
            # 1: CODIGO INRED, 3: DEPARTAMENTO, 5: CENTRO POBLADO, 6: LAT, 7: LON
            mapping[code] = {
                "department": str(row[3]).strip() if row[3] else "",
                "centro_poblado": str(row[5]).strip() if row[5] else "",
                "latitude": float(row[6] or 0),
                "longitude": float(row[7] or 0),
                "estado": "Implementación",
                "fecha_inicio": "",
            }
        wb.close()
        logger.info(f"Mapeo Excel cargado: {len(mapping)} codigos INRED")
    except Exception as e:
        logger.warning(f"No se pudo cargar Excel de mapeo: {e}")

    return mapping


def load_operational_sites(filepath: str = "excel/base_operacion.xlsx") -> Dict:
    """Carga el Excel de juntas en operacion y retorna mapeo por codigo INRED.
    Retorna dict: codigo_inred -> {department, fecha_inicio, estado}
    """
    ops = {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            code = str(row[1]).strip()
            dept = str(row[3]).strip() if row[3] else ""
            fecha = row[14]
            fecha_str = str(fecha)[:10] if fecha else ""

            ops[code] = {
                "department": dept,
                "fecha_inicio": fecha_str,
                "estado": "Operación",
            }
        wb.close()
        logger.info(f"Juntas en operacion cargadas: {len(ops)}")
    except Exception as e:
        logger.warning(f"No se pudo cargar Excel de operacion: {e}")

    return ops


def load_group_mapping(filepath: str = "excel/Grupo 1 - Juntas de Internet.xlsx") -> Dict:
    """Carga el archivo Excel de grupos de juntas y genera mapeo por codigo INRED.
    Retorna dict: codigo_inred -> grupo (ej. "Grupo 1").
    """
    grupos = {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        header = [str(c).strip().upper() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        # La columna del código INRED puede variar segun el Excel; localizarla por nombre.
        def _find_col(*names):
            for i, h in enumerate(header):
                if any(n in h for n in names):
                    return i
            return None

        code_col = _find_col("CODIGO INRED", "CODIGO", "CÓDIGO")
        if code_col is None:
            raise ValueError(f"Columna de código INRED no encontrada en {filepath} (headers: {header})")

        # Inferir nombre del grupo (ej. "Grupo 1") desde el nombre del archivo
        import re as _re
        m = _re.search(r"Grupo\s*\d+", filepath, _re.IGNORECASE)
        grupo_name = m.group(0).strip() if m else "Grupo 1"

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or code_col >= len(row) or not row[code_col]:
                continue
            code = str(row[code_col]).strip()
            if not code:
                continue
            grupos[code] = grupo_name
        wb.close()
        logger.info(f"Grupos de juntas cargados: {len(grupos)} juntas en {grupo_name}")
    except Exception as e:
        logger.warning(f"No se pudo cargar Excel de grupos: {e}")

    return grupos

# ============================================================================
# ============================================================================
# DATACLASSES
# ============================================================================
@dataclass
class SiteData:
    """Datos principales de un Site (AP)"""
    site_id: str
    site_name: str
    site_status: str
    location: Dict
    elevation: float
    device_count: int
    device_outage_count: int
    download_capacity: float
    upload_capacity: float
    last_seen: str
    sla: float
    ap_count: int = 0
    ap_online: int = 0
    lb5_count: int = 0
    lb5_online: int = 0
    department: Optional[str] = None
    estado: str = "Implementación"
    fecha_inicio: str = ""
    contact_email: str = ""
    inred_code: Optional[str] = None
    grupo: str = ""
    
@dataclass
class DeviceData:
    """Datos de un Device (Hogar/Cliente)"""
    device_id: str
    device_name: str
    device_model: str
    device_status: str
    ip_address: str
    mac_address: str
    site_id: str
    site_name: str
    signal_strength: float
    traffic_summary: float
    rx_capacity: float
    tx_capacity: float
    rx_throughput: float
    tx_throughput: float
    downlink_utilization: float
    uplink_utilization: float
    last_seen: str
    uptime: int
    temperature: Optional[float] = None
    cpu_usage: Optional[float] = None
    ram_usage: Optional[float] = None

# ============================================================================
# CLIENTE API UBIQUITI
# ============================================================================
class UbiquitiAPIClient:
    """Cliente para interactuar con la API de Ubiquiti"""
    
    def __init__(self, base_url: str, token: str, verify_ssl: bool = True):
        self.base_url = base_url
        self.token = token
        self.verify_ssl = verify_ssl
        self.headers = {
            "accept": "application/json",
            "x-auth-token": token,
            "Content-Type": "application/json"
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        self.session.verify = verify_ssl
    
    def _make_request(self, endpoint: str) -> Dict:
        """Realiza una solicitud GET a la API"""
        url = f"{self.base_url}{endpoint}"
        try:
            response = self.session.get(url, timeout=60)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Error en API request {endpoint}: {e}")
            return {}
    
    def get_sites(self) -> List[Dict]:
        """Obtiene todos los sites (APs)"""
        logger.info("Obteniendo sites...")
        return self._make_request("/sites") or []
    
    def get_devices(self) -> List[Dict]:
        logger.info("Obteniendo devices...")
        return self._make_request("/devices") or []

# ============================================================================
# PROCESADOR DE DATOS
# ============================================================================
class DataProcessor:
    """Procesa datos de la API en objetos estructurados"""
    
    def __init__(self, department_mapping: Dict = None):
        self.department_mapping = department_mapping or {}
    
    @staticmethod
    def safe_float(value, default: float = 0.0) -> float:
        """Convierte un valor a float de forma segura, manejando None"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    @staticmethod
    def safe_int(value, default: int = 0) -> int:
        """Convierte un valor a int de forma segura, manejando None"""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    def get_site_mapping(self, site_id: str, site_name: str, ip_address: str) -> Tuple[Dict, Optional[str]]:
        """Obtiene el mapeo completo y el código INRED encontrado"""
        empty = {"department": None, "estado": "Implementación", "fecha_inicio": "", "grupo": ""}
        
        if site_id in self.department_mapping:
            return self.department_mapping[site_id], site_id
        if ip_address in self.department_mapping:
            return self.department_mapping[ip_address], ip_address

        # Intentar encontrar el código INRED en cualquier parte del nombre del sitio
        for code in self.department_mapping.keys():
            if code and str(code) in site_name:
                return self.department_mapping[code], code


        if "-" in site_name:
            parts = site_name.split("-")
            if len(parts) >= 2:
                empty["department"] = parts[1].strip()

        return empty, None
    
    def process_site(self, site_data: Dict) -> Optional[SiteData]:
        """Procesa datos de un site"""
        try:
            # PROTECCIÓN 1: Usar 'or {}' garantiza que si la API devuelve 'null', 
            # Python asigne un diccionario vacío en lugar de None
            identification = site_data.get("identification") or {}
            description = site_data.get("description") or {}
            overview = site_data.get("overview") or {}
            
            site_id = identification.get("id", "")
            site_name = identification.get("name", "unknown")
            site_status = identification.get("status", "unknown")
            
            # Ubicación
            location = description.get("location") or {}
            
            # PROTECCIÓN 2: Usar tus funciones self.safe_float y self.safe_int
            download_capacity = self.safe_float(overview.get("downlinkCapacity"))
            upload_capacity = self.safe_float(overview.get("uplinkCapacity"))
            
            # Dispositivos
            device_count = self.safe_int(description.get("deviceCount"))
            device_outage_count = self.safe_int(description.get("deviceOutageCount"))
            
            # Contacto
            contact = description.get("contact") or {}
            contact_email = contact.get("email", "")
            
            # SLA
            sla = self.safe_float(description.get("sla"))
            
            # Obtener zona y departamento protegiendo la lista de IPs
            ip_addresses = description.get("ipAddresses") or [""]
            ip_address = ip_addresses[0] if ip_addresses else ""
            mapping, inred_code = self.get_site_mapping(site_id, site_name, ip_address)
            department = mapping.get("department")
            estado = mapping.get("estado", "Implementación")
            fecha_inicio = mapping.get("fecha_inicio", "")
            grupo = mapping.get("grupo", "")


            site_obj = SiteData(
                site_id=site_id,
                site_name=site_name,
                site_status=site_status,
                location=location,
                elevation=self.safe_float(description.get("elevation")),
                device_count=device_count,
                device_outage_count=device_outage_count,
                download_capacity=download_capacity,
                upload_capacity=upload_capacity,
                last_seen=overview.get("lastSeen", ""),
                sla=sla,
                department=department,
                estado=estado,
                fecha_inicio=fecha_inicio,
                contact_email=contact_email,
                inred_code=inred_code,
                grupo=grupo
            )
            
            return site_obj

        except Exception as e:
            logger.error(f"Error procesando site: {e}")
            return None
    
    def process_device(self, device_data: Dict) -> Optional[DeviceData]:
        """Procesa datos de un device"""
        try:
            # PROTECCIÓN 1
            identification = device_data.get("identification") or {}
            overview = device_data.get("overview") or {}
            
            device_id = identification.get("id", "")
            device_name = identification.get("name", "unknown")
            device_model = identification.get("model", "unknown")
            device_status = overview.get("status") or identification.get("status", "unknown")
            # --- CAMBIO: Buscamos en 'ipAddress' o 'ipAddressList' ---
            ip_address = device_data.get("ipAddress") or (device_data.get("ipAddressList", [])[0] if device_data.get("ipAddressList") else "")
            mac_address = identification.get("mac", "")
            
            # Site padre
            site_info = identification.get("site") or {}
            site_id = site_info.get("id", "")
            site_name = site_info.get("name", "unknown")
            
            # PROTECCIÓN 2: Reemplazo masivo de float() por self.safe_float()
            downlink = self.safe_float(overview.get("downlinkCapacity"))
            uplink = self.safe_float(overview.get("uplinkCapacity"))
            
            # El tráfico real es la utilización (throughput en bps/Bps según dispositivo)
            # Solo si el dispositivo está activo reportamos tráfico, evitando datos stale de la API
            is_active = str(device_status).lower() == "active"
            
            if is_active:
                downlink_util = self.safe_float(overview.get("downlinkUtilization"))
                uplink_util = self.safe_float(overview.get("uplinkUtilization"))
            else:
                downlink_util = 0.0
                uplink_util = 0.0

            traffic_summary = downlink_util + uplink_util
            rx_capacity = downlink
            tx_capacity = uplink
            rx_throughput = downlink_util
            tx_throughput = uplink_util
            
            # Calidad de señal (con valor por defecto -100)
            signal_strength = self.safe_float(overview.get("signal"), -100.0)
            
            # Uso de recursos
            cpu_usage = self.safe_float(overview.get("cpu"))
            ram_usage = self.safe_float(overview.get("ram"))
            temperature = self.safe_float(overview.get("temperature"))
            uptime = self.safe_int(overview.get("uptime"))
            
            device_obj = DeviceData(
                device_id=device_id,
                device_name=device_name,
                device_model=device_model,
                device_status=device_status,
                ip_address=ip_address,
                mac_address=mac_address,
                site_id=site_id,
                site_name=site_name,
                signal_strength=signal_strength,
                traffic_summary=traffic_summary,
                rx_capacity=rx_capacity,
                tx_capacity=tx_capacity,
                rx_throughput=rx_throughput,
                tx_throughput=tx_throughput,
                downlink_utilization=downlink_util,
                uplink_utilization=uplink_util,
                last_seen=overview.get("lastSeen", datetime.now().isoformat()),
                uptime=uptime,
                temperature=temperature,
                cpu_usage=cpu_usage,
                ram_usage=ram_usage
            )
            return device_obj
        except Exception as e:
            logger.error(f"Error procesando device: {e}")
            return None
# ============================================================================
# GESTOR DE INFLUXDB
# ============================================================================
class InfluxDBManager:
    """Gestiona la conexión y escritura de datos en InfluxDB"""
    
    def __init__(self, host: str, port: int, user: str, password: str, database: str):
        self.client = InfluxDBClient(
            host=host,
            port=port,
            username=user,
            password=password,
            database=database,
            timeout=5
        )
        self.database = database
        self._ensure_database()
        self._ensure_retention_policies()
    
    def _ensure_database(self):
        """Crea la base de datos si no existe"""
        try:
            databases = self.client.get_list_database()
            if self.database not in [db['name'] for db in databases]:
                self.client.create_database(self.database)
                logger.info(f"Base de datos '{self.database}' creada")
        except Exception as e:
            logger.error(f"Error verificando/creando base de datos: {e}")

    def _ensure_retention_policies(self):
        """Garantiza las retention policies de la base de datos:
        - 'autogen' (default) con retencion infinita para sites/devices/dept_summary.
        - 'traffic_90d' con retencion de RETENTION_DAYS dias solo para la medicion traffic.
        """
        try:
            rps = self.client.get_list_retention_policies(database=self.database)
            rp_names = {rp["name"] for rp in rps}

            if "autogen" in rp_names:
                self.client.query(
                    f"ALTER RETENTION POLICY autogen ON {self.database} DURATION 0s"
                )
            else:
                self.client.create_retention_policy(
                    "autogen", "0s", 1, self.database, default=True
                )

            traffic_rp = f"traffic_{RETENTION_DAYS}d"
            if traffic_rp not in rp_names:
                self.client.create_retention_policy(
                    traffic_rp, f"{RETENTION_DAYS}d", 1, self.database, default=False
                )
                logger.info(f"Retention policy '{traffic_rp}' ({RETENTION_DAYS}d) creada")
            else:
                logger.info(f"Retention policy '{traffic_rp}' ya existe")
        except Exception as e:
            logger.error(f"Error verificando/creando retention policies: {e}")
    
    def write_site_data(self, sites: List[SiteData]) -> int:
        """Escribe datos de sites en InfluxDB"""
        points = []
        
        for site in sites:
            point = {
                "measurement": "sites",
                "tags": {
                    "site_id": site.site_id,
                    "site_name": site.site_name,
                    "status": site.site_status,
                    "department": site.department or "unknown",
                    "estado": site.estado or "Implementación",
                    "fecha_inicio": site.fecha_inicio or "",
                    "inred_code": site.inred_code or "unknown",
                    "grupo": site.grupo or "Sin grupo",
                },
                "fields": {
                    "device_count": site.device_count,
                    "device_outage_count": site.device_outage_count,
                    "devices_available": site.device_count - site.device_outage_count,
                    "ap_count": site.ap_count,
                    "ap_online": site.ap_online,
                    "lb5_count": site.lb5_count,
                    "lb5_online": site.lb5_online,
                    "download_capacity": float(site.download_capacity),
                    "upload_capacity": float(site.upload_capacity),
                    "elevation": float(site.elevation),
                    "sla": float(site.sla),
                    # AQUI ESTABA EL ERROR (0 -> 0.0)
                    "latitude": float(site.location.get("latitude", 0.0)),
                    "longitude": float(site.location.get("longitude", 0.0)),
                    "last_seen": site.last_seen
                },
                # Se escribe con la hora del ciclo (no last_seen): un site caido
                # hace >24h tiene last_seen viejo y con ese timestamp su punto
                # quedaria fuera de la ventana de consulta (24h), haciendolo
                # desaparecer del dashboard y de /api/alerts (el notificador
                # Telegram lo daria por "recuperado" falsamente).
                "time": int(time.time() * 1e9)
            }
            points.append(point)
        
        try:
            self.client.write_points(points, time_precision='n')
            logger.info(f"Escritos {len(points)} registros de sites")
            return len(points)
        except Exception as e:
            logger.error(f"Error escribiendo sites: {e}")
            return 0
     
    def write_device_data(self, devices: List[DeviceData]) -> int:
        """Escribe datos de devices en InfluxDB"""
        points = []
        
        for device in devices:
            point = {
                "measurement": "devices",
                "tags": {
                    "device_id": device.device_id,
                    "device_name": device.device_name,
                    "device_model": device.device_model,
                    "status": device.device_status,
                    "site_id": device.site_id,
                    "site_name": device.site_name,
                    "ip_address": device.ip_address,
                    "mac_address": device.mac_address
                },
                "fields": {
                    "signal_strength": float(device.signal_strength),
                    "traffic_summary": float(device.traffic_summary),
                    "rx_capacity": float(device.rx_capacity),
                    "tx_capacity": float(device.tx_capacity),
                    "rx_throughput": float(device.rx_throughput),
                    "tx_throughput": float(device.tx_throughput),
                    "downlink_utilization": float(device.downlink_utilization),
                    "uplink_utilization": float(device.uplink_utilization),
                    "uptime": int(device.uptime),
                    "temperature": float(device.temperature or 0.0),
                    "cpu_usage": float(device.cpu_usage or 0.0),
                    "ram_usage": float(device.ram_usage or 0.0),
                    "last_seen": device.last_seen
                },
                "time": self._parse_timestamp(device.last_seen) or int(time.time() * 1e9)
            }
            points.append(point)
        
        try:
            self.client.write_points(points, time_precision='n')
            logger.info(f"Escritos {len(points)} registros de devices")
            return len(points)
        except Exception as e:
            logger.error(f"Error escribiendo devices: {e}")
            return 0
    
    def write_traffic_data(self, sites: List[SiteData]) -> int:
        """Escribe el trafico Rx/Tx por site en la medicion 'traffic',
        con su propia retention policy de RETENTION_DAYS dias, dejando
        la medicion 'sites' solo con el estado (retencion infinita).
        Los points mas antiguos que la retencion se descartan para no
        provocar "partial write: points beyond retention".
        """
        cutoff = time.time() - (RETENTION_DAYS * 24 * 3600)
        points = []

        for site in sites:
            ts = self._parse_timestamp(site.last_seen) or int(time.time() * 1e9)
            if ts / 1e9 < cutoff:
                continue

            point = {
                "measurement": "traffic",
                "tags": {
                    "site_id": site.site_id,
                    "site_name": site.site_name,
                    "department": site.department or "unknown",
                    "estado": site.estado or "Implementación",
                    "grupo": site.grupo or "Sin grupo",
                },
                "fields": {
                    "download_capacity": float(site.download_capacity),
                    "upload_capacity": float(site.upload_capacity),
                },
                "time": ts,
            }
            points.append(point)

        try:
            if not points:
                logger.info("Sin puntos de trafico para escribir (todos fuera de retencion)")
                return 0
            self.client.write_points(points, retention_policy=f"traffic_{RETENTION_DAYS}d", time_precision='n')
            logger.info(f"Escritos {len(points)} registros de trafico")
            return len(points)
        except Exception as e:
            logger.error(f"Error escribiendo trafico: {e}")
            return 0

    def write_aggregated_data(self, sites: List[SiteData]) -> int:
        """Escribe datos agregados por departamento"""
        points = []

        # Agrupar por departamento
        dept_summary = {}
        for site in sites:
            key = site.department or "unknown"
            if key not in dept_summary:
                dept_summary[key] = {
                    "site_count": 0,
                    "total_devices": 0,
                    "active_devices": 0,
                    "failed_devices": 0,
                    "total_download": 0,
                    "total_upload": 0,
                    "average_sla": 0,
                    "sla_count": 0
                }
            
            dept_summary[key]["site_count"] += 1
            dept_summary[key]["total_devices"] += site.device_count
            dept_summary[key]["active_devices"] += (site.device_count - site.device_outage_count)
            dept_summary[key]["failed_devices"] += site.device_outage_count
            dept_summary[key]["total_download"] += site.download_capacity
            dept_summary[key]["total_upload"] += site.upload_capacity
            dept_summary[key]["average_sla"] += site.sla
            dept_summary[key]["sla_count"] += 1
        
        # Crear puntos de datos agregados
        for department, summary in dept_summary.items():
            average_sla = summary["average_sla"] / summary["sla_count"] if summary["sla_count"] > 0 else 0
            
            point = {
                "measurement": "dept_summary",
                "tags": {
                    "department": department
                },
              "fields": {
                    "site_count": summary["site_count"],
                    "total_devices": summary["total_devices"],
                    "active_devices": summary["active_devices"],
                    "failed_devices": summary["failed_devices"],
                    # AQUI ESTABA EL ERROR (0 -> 0.0)
                    "device_availability_percent": float((summary["active_devices"] / summary["total_devices"] * 100) if summary["total_devices"] > 0 else 0.0),
                    "total_download": float(summary["total_download"]),
                    "total_upload": float(summary["total_upload"]),
                    "average_sla": float(average_sla)
                },
                "time": int(datetime.now(pytz.UTC).timestamp() * 1e9)
            }
            points.append(point)
        
        try:
            self.client.write_points(points, time_precision='n')
            logger.info(f"Escritos {len(points)} registros agregados")
            return len(points)
        except Exception as e:
            logger.error(f"Error escribiendo datos agregados: {e}")
            return 0
    
    @staticmethod
    def _parse_timestamp(timestamp_str: str) -> Optional[int]:
        """Convierte timestamp ISO a nanosegundos (UTC), retorna None si no es valido"""
        try:
            if isinstance(timestamp_str, str) and timestamp_str.strip():
                ts = timestamp_str.replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=pytz.UTC)
                return int(dt.timestamp() * 1e9)
            return None
        except Exception:
            return None

# ============================================================================
# MONITOR PRINCIPAL
# ============================================================================
class UbiquitiMonitor:
    """Orquestador principal del sistema de monitoreo"""
    
    def __init__(self, config_dict: Dict):
        self.config = config_dict
        ubiq = config_dict["ubiquiti"]
        influx = config_dict["influxdb"]

        self.api_client = UbiquitiAPIClient(ubiq["api_url"], ubiq["token"], ubiq["verify_ssl"])
        self.data_processor = DataProcessor(self.config["departments"])
        self.influx_manager = InfluxDBManager(
            influx["host"],
            influx["port"],
            influx["username"],
            influx["password"],
            influx["database"]
        )

    def fetch_and_process(self):
        """Obtiene y procesa datos de la API"""
        logger.info("=" * 60)
        logger.info("Iniciando ciclo de recolección de datos")
        logger.info("=" * 60)
        
        # Obtener datos de la API
        raw_sites = self.api_client.get_sites()
        raw_devices = self.api_client.get_devices()
        
        # Procesar sites
        sites = []
        sites_map = {}
        
        for raw_site in raw_sites:
            site = self.data_processor.process_site(raw_site)
            if site:
                sites.append(site)
                sites_map[site.site_id] = site
        
        logger.info(f"Ciclo completado: {len(sites)} sites reales")
        
# Procesar devices
        devices = []
        for raw_device in raw_devices:
            device = self.data_processor.process_device(raw_device)
            if device:
                devices.append(device)

        # Recomputar tráfico por site desde utilización de devices
        for site in sites:
            site.download_capacity = 0.0
            site.upload_capacity = 0.0
            site.ap_count = 0
            site.ap_online = 0
            site.lb5_count = 0
            site.lb5_online = 0
        for device in devices:
            site = sites_map.get(device.site_id)
            if site:
                # Throughput real (bps) = capacidad del enlace × fracción de uso activo
                site.download_capacity += device.downlink_utilization * device.rx_capacity
                site.upload_capacity += device.uplink_utilization * device.tx_capacity
                # Contar APs (modelos R5AC-Lite, etc.)
                model = (device.device_model or "").strip()
                is_active = (device.device_status or "").lower() == "active"
                if model in ("R5AC-Lite", "LAP-120", "LAP-GPS"):
                    site.ap_count += 1
                    if is_active:
                        site.ap_online += 1
                elif model == "LB5":
                    site.lb5_count += 1
                    if is_active:
                        site.lb5_online += 1

        logger.info(f"Procesados {len(devices)} devices")
        
        return sites, devices
    
    def write_to_influx(self, sites: List[SiteData], devices: List[DeviceData]) -> bool:
        """Escribe datos a InfluxDB"""
        try:
            # Escribir datos de sites
            sites_count = self.influx_manager.write_site_data(sites)
            
            # Escribir datos de devices
            devices_count = self.influx_manager.write_device_data(devices)
            
            # Escribir datos de trafico (retention separada)
            traffic_count = self.influx_manager.write_traffic_data(sites)
            
            # Escribir datos agregados
            aggregated_count = self.influx_manager.write_aggregated_data(sites)
            
            logger.info(f"Total escritas: {sites_count} sites + {devices_count} devices + {traffic_count} trafico + {aggregated_count} agregados")
            return True
        except Exception as e:
            logger.error(f"Error escribiendo en InfluxDB: {e}")
            return False
    
    def run(self):
        """Ejecuta el monitor en forma continua"""
        logger.info("Iniciando monitor de Ubiquiti...")
        
        cycle = 0
        while True:
            try:
                cycle += 1
                logger.info(f"\n>>> Ciclo #{cycle} - {datetime.now().isoformat()}")
                
                # Obtener y procesar datos
                sites, devices = self.fetch_and_process()
                
                if sites or devices:
                    # Escribir a InfluxDB
                    self.write_to_influx(sites, devices)
                    
                    # Resumen
                    logger.info(f"Resumen: {len(sites)} sites, {len(devices)} devices")
                    
                    # Mostrar algunos detalles
                    for site in sites[:3]:  # Mostrar primeros 3 sites
                        logger.info(
                            f"  Site: {site.site_name} | Status: {site.site_status} | "
                            f"Devices: {site.device_count} ({site.device_outage_count} caídos) | "
                            f"Dept: {site.department}"
                        )
                else:
                    logger.warning("No se obtuvieron datos de la API")
                
                logger.info(f"Próxima ejecución en {self.config['monitoring']['polling_interval']}s")
                time.sleep(self.config['monitoring']['polling_interval'])
                
            except KeyboardInterrupt:
                logger.info("Monitor detenido por el usuario")
                break
            except Exception as e:
                logger.error(f"Error en ciclo de monitoreo: {e}", exc_info=True)
                logger.info(f"Reintentando en {self.config['monitoring']['polling_interval']}s...")
                time.sleep(self.config['monitoring']['polling_interval'])

# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================
if __name__ == "__main__":
    # Cargar mapeo desde Excel de instalaciones (carpeta excel/)
    excel_mapping = load_excel_mapping("excel/base_instalaciones.xlsx")
    operaciones = load_operational_sites("excel/base_operacion.xlsx")
    grupos = load_group_mapping("excel/Grupo 1 - Juntas de Internet.xlsx")

    # Merge: priorizar info de operacion pero rellenar departamentos faltantes desde instalaciones
    final_mapping = excel_mapping.copy()
    
    for code, op_data in operaciones.items():
        if code in final_mapping:
            # Actualizar estado y fecha de inicio
            final_mapping[code]["estado"] = op_data["estado"]
            final_mapping[code]["fecha_inicio"] = op_data.get("fecha_inicio", "")
            
            # Si el departamento en instalaciones está vacío, usar el de operaciones
            if not final_mapping[code].get("department") and op_data.get("department"):
                final_mapping[code]["department"] = op_data["department"]
        else:
            # Junta que solo existe en base_operacion
            # Intentar asignar departamento si existe en op_data
            final_mapping[code] = {
                "department": op_data.get("department", ""),
                "centro_poblado": "",
                "latitude": 0,
                "longitude": 0,
                "estado": op_data["estado"],
                "fecha_inicio": op_data.get("fecha_inicio", ""),
            }
    
    # Asignar grupo a las juntas del Excel "Grupo N - Juntas de Internet.xlsx"
    for code, grupo in grupos.items():
        if code in final_mapping:
            final_mapping[code]["grupo"] = grupo
        else:
            final_mapping[code] = {
                "department": "",
                "centro_poblado": "",
                "latitude": 0,
                "longitude": 0,
                "estado": "Implementación",
                "fecha_inicio": "",
                "grupo": grupo,
            }
    
    # Inyectar el mapeo final en la configuración global
    GlobalConfig["departments"] = final_mapping
    
    # Crear y ejecutar monitor
    monitor = UbiquitiMonitor(GlobalConfig)
    monitor.run()
